"""
Convert consolidated CSV from Abaqus extraction into plot-ready Excel workbooks.

Structure:
  - Summary sheet:  peak values per step per variable
  - Per variable:   separate sheets for Loading & Relaxation phases
                    1D sweep => rows = X positions, columns = step names
                    2D sweep => stacked Y x X blocks per step
  - AllSteps sheets: all steps in one sheet (Loading color-coded vs Relaxation)

Steps are classified automatically:
  - Loading:     odd steps (1st, 3rd, 5th, ...)
  - Relaxation:  even steps (2nd, 4th, 6th, ...)

Usage:
    python format_to_excel.py                            # auto-finds *_consolidated.csv
    python format_to_excel.py  ECM_consolidated.csv      # specific file
"""

import os
import sys
import csv
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ── Style definitions ───────────────────────────────────────────────────────
HEADER_FILL_DARK = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=10, color="FFFFFF")
LOADING_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
LOADING_FONT = Font(bold=True, size=10, color="90CDF4")
RELAX_FILL = PatternFill(start_color="322659", end_color="322659", fill_type="solid")
RELAX_FONT = Font(bold=True, size=10, color="D6BCFA")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
POS_LABEL_FONT = Font(bold=False, size=10, color="A0AEC0")
POS_LABEL_FILL = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")
SUMMARY_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")


def _style_header_row(ws, row_num, num_cols, fill=None, font=None):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = font or HEADER_FONT_WHITE
        cell.fill = fill or HEADER_FILL_DARK
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _write_value(ws, row_num, col_num, val_str):
    try:
        val = float(val_str)
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.number_format = '0.0000E+00'
    except (ValueError, TypeError):
        cell = ws.cell(row=row_num, column=col_num, value=val_str)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")


def _write_pos_label(ws, row_num, col_num, value):
    cell = ws.cell(row=row_num, column=col_num, value=value)
    cell.font = POS_LABEL_FONT
    cell.fill = POS_LABEL_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")


def _auto_fit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            if cell.value is not None and not isinstance(cell, type(None)):
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 3, 22)


def _safe_sheet_name(name, max_len=31):
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, '')
    return name[:max_len]


# ── Parse the consolidated CSV ──────────────────────────────────────────────
def parse_csv(csv_path):
    """Return metadata dict, header list, and data rows."""
    meta = {}
    header = None
    rows = []

    with open(csv_path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("#"):
                content = line.lstrip("# ").strip()
                if ":" in content:
                    key, val = content.split(":", 1)
                    meta[key.strip()] = val.strip()
                continue
            if header is None:
                header = [h.strip() for h in line.split(",")]
                continue
            values = [v.strip() for v in line.split(",")]
            row = {}
            for i, h in enumerate(header):
                row[h] = values[i] if i < len(values) else ""
            rows.append(row)

    return meta, header, rows


def detect_variable_labels(header):
    """
    Extract base variable labels from header.
    e.g. header has LE22_Max, LE22_Min, S22_Max, S22_Min
    returns ['LE22', 'S22'] in order
    """
    var_start = header.index("YLabel") + 1
    raw_cols = header[var_start:]
    labels = []
    seen = set()
    for col in raw_cols:
        # Strip _Max or _Min suffix
        if col.endswith("_Max"):
            base = col[:-4]
        elif col.endswith("_Min"):
            base = col[:-4]
        else:
            base = col
        if base not in seen:
            seen.add(base)
            labels.append(base)
    return labels


def build_data(rows, var_labels):
    """
    Build structured data dict.

    Returns:
        x_labels:   sorted unique X position labels
        y_labels:   sorted unique Y position labels
        step_order: list of step names in order of appearance
        data:       {(step, xi, yi): {var_label: (max_val, min_val)}}
    """
    x_set = set()
    y_set = set()
    step_order = []
    step_seen = set()
    data = {}

    for row in rows:
        step = row.get("Step", "")
        xi = int(row.get("Xi", 0))
        yi = int(row.get("Yi", 0))
        x_label = row.get("XLabel", "")
        y_label = row.get("YLabel", "")

        x_set.add((xi, x_label))
        y_set.add((yi, y_label))

        if step not in step_seen:
            step_seen.add(step)
            step_order.append(step)

        key = (step, xi, yi)
        if key not in data:
            data[key] = {}

        for label in var_labels:
            max_col = "%s_Max" % label
            min_col = "%s_Min" % label
            max_val = row.get(max_col, "N/A")
            min_val = row.get(min_col, "N/A")
            data[key][label] = (max_val, min_val)

    # Sort by index
    x_labels = [lbl for _, lbl in sorted(x_set)]
    y_labels = [lbl for _, lbl in sorted(y_set)]

    return x_labels, y_labels, step_order, data


# ── Write Summary sheet ─────────────────────────────────────────────────────
def write_summary(wb, set_name, var_labels, x_labels, y_labels, step_order,
                  loading_steps, relaxation_steps, data):
    ws = wb.create_sheet(title="Summary")

    ws.cell(row=1, column=1, value="Element Set: %s" % set_name)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = HEADER_FILL_DARK

    ws.cell(row=2, column=1,
            value="Steps: %d  (%d Loading + %d Relaxation)" % (
                len(step_order), len(loading_steps), len(relaxation_steps)))
    ws.cell(row=2, column=1).font = Font(size=10, color="A0AEC0")

    ws.cell(row=3, column=1,
            value="Regions: %d  (%d X x %d Y)" % (
                len(x_labels) * max(len(y_labels), 1),
                len(x_labels), max(len(y_labels), 1)))
    ws.cell(row=3, column=1).font = Font(size=10, color="A0AEC0")

    # Peak-values table
    sum_row = 5
    ws.cell(row=sum_row, column=1, value="Step")
    ws.cell(row=sum_row, column=2, value="Phase")
    col_idx = 3
    for label in var_labels:
        ws.cell(row=sum_row, column=col_idx, value="%s Max" % label)
        ws.cell(row=sum_row, column=col_idx + 1, value="%s Min" % label)
        col_idx += 2
    _style_header_row(ws, sum_row, col_idx - 1, SUMMARY_HEADER_FILL, SUMMARY_HEADER_FONT)

    for step_idx, step in enumerate(step_order):
        r = sum_row + step_idx + 1
        phase = "Loading" if step in loading_steps else "Relaxation"
        ws.cell(row=r, column=1, value=step).border = THIN_BORDER
        phase_cell = ws.cell(row=r, column=2, value=phase)
        phase_cell.border = THIN_BORDER
        if phase == "Loading":
            phase_cell.font = LOADING_FONT
            phase_cell.fill = LOADING_FILL
        else:
            phase_cell.font = RELAX_FONT
            phase_cell.fill = RELAX_FILL

        col_idx = 3
        for label in var_labels:
            peak_max = None
            peak_min = None
            for xi in range(len(x_labels)):
                for yi in range(len(y_labels)):
                    key = (step, xi, yi)
                    if key in data and label in data[key]:
                        try:
                            mv = float(data[key][label][0])
                            if peak_max is None or mv > peak_max:
                                peak_max = mv
                        except (ValueError, TypeError):
                            pass
                        try:
                            nv = float(data[key][label][1])
                            if peak_min is None or nv < peak_min:
                                peak_min = nv
                        except (ValueError, TypeError):
                            pass

            if peak_max is not None:
                _write_value(ws, r, col_idx, "%.8e" % peak_max)
            else:
                ws.cell(row=r, column=col_idx, value="N/A").border = THIN_BORDER
            if peak_min is not None:
                _write_value(ws, r, col_idx + 1, "%.8e" % peak_min)
            else:
                ws.cell(row=r, column=col_idx + 1, value="N/A").border = THIN_BORDER
            col_idx += 2

    _auto_fit(ws)


# ── Write per-variable data sheets ──────────────────────────────────────────
def write_variable_sheets(wb, var_labels, x_labels, y_labels, step_order,
                          loading_steps, relaxation_steps, data, is_2d, set_name):
    phases = [
        ("Loading", loading_steps, LOADING_FILL, LOADING_FONT),
        ("Relax", relaxation_steps, RELAX_FILL, RELAX_FONT),
    ]

    pos_labels = x_labels if len(x_labels) > 1 else y_labels
    pos_axis = "X" if len(x_labels) > 1 else "Y"

    for label in var_labels:
        for stat_type, stat_idx in [("Max", 0), ("Min", 1)]:
            # ── Phase-specific sheets (Loading / Relaxation) ──
            for phase_name, phase_steps, phase_fill, phase_font in phases:
                if not phase_steps:
                    continue

                sheet_name = _safe_sheet_name("%s_%s_%s" % (label, stat_type, phase_name))
                ws = wb.create_sheet(title=sheet_name)

                if is_2d:
                    _write_2d_sheet(ws, label, stat_type, stat_idx, phase_name,
                                   phase_steps, phase_fill, phase_font,
                                   x_labels, y_labels, data)
                else:
                    _write_1d_sheet(ws, label, stat_type, stat_idx, phase_name,
                                   phase_steps, phase_fill, phase_font,
                                   pos_labels, pos_axis, data, set_name)
                _auto_fit(ws)

            # ── AllSteps sheet ──
            if len(step_order) > 1:
                sheet_name = _safe_sheet_name("%s_%s_AllSteps" % (label, stat_type))
                ws = wb.create_sheet(title=sheet_name)

                if is_2d:
                    current_row = 1
                    for step in step_order:
                        phase = "Loading" if step in loading_steps else "Relaxation"
                        pf = LOADING_FILL if phase == "Loading" else RELAX_FILL
                        pfont = LOADING_FONT if phase == "Loading" else RELAX_FONT

                        ws.cell(row=current_row, column=1,
                                value="%s  |  %s %s  |  %s" % (step, label, stat_type, phase))
                        ws.cell(row=current_row, column=1).font = pfont
                        ws.cell(row=current_row, column=1).fill = pf
                        ws.merge_cells(start_row=current_row, start_column=1,
                                       end_row=current_row, end_column=len(x_labels) + 1)
                        current_row += 1

                        ws.cell(row=current_row, column=1, value="Y \\ X")
                        for ci, xl in enumerate(x_labels):
                            ws.cell(row=current_row, column=ci + 2, value=xl)
                        _style_header_row(ws, current_row, len(x_labels) + 1)
                        current_row += 1

                        for yi_idx, yl in enumerate(y_labels):
                            _write_pos_label(ws, current_row, 1, yl)
                            for xi_idx in range(len(x_labels)):
                                key = (step, xi_idx, yi_idx)
                                if key in data and label in data[key]:
                                    _write_value(ws, current_row, xi_idx + 2,
                                                 data[key][label][stat_idx])
                            current_row += 1
                        current_row += 1  # blank separator row

                else:
                    # 1D: rows=positions, columns=ALL step names (color-coded)
                    ws.cell(row=1, column=1,
                            value="%s  |  %s %s  |  All Steps" % (set_name, label, stat_type))
                    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
                    ws.merge_cells(start_row=1, start_column=1,
                                   end_row=1, end_column=len(step_order) + 1)
                    ws.cell(row=1, column=1).fill = HEADER_FILL_DARK

                    ws.cell(row=2, column=1, value="Position_%s (mm)" % pos_axis)
                    for si, step in enumerate(step_order):
                        cell = ws.cell(row=2, column=si + 2, value=step)
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(horizontal="center")
                        if step in loading_steps:
                            cell.font = LOADING_FONT
                            cell.fill = LOADING_FILL
                        else:
                            cell.font = RELAX_FONT
                            cell.fill = RELAX_FILL

                    pos_cell = ws.cell(row=2, column=1)
                    pos_cell.font = HEADER_FONT_WHITE
                    pos_cell.fill = HEADER_FILL_DARK
                    pos_cell.border = THIN_BORDER

                    for ci, pl in enumerate(pos_labels):
                        row_num = ci + 3
                        _write_pos_label(ws, row_num, 1, pl)
                        for si, step in enumerate(step_order):
                            if pos_axis == "X":
                                key = (step, ci, 0)
                            else:
                                key = (step, 0, ci)
                            if key in data and label in data[key]:
                                _write_value(ws, row_num, si + 2,
                                             data[key][label][stat_idx])

                _auto_fit(ws)


def _write_1d_sheet(ws, label, stat_type, stat_idx, phase_name,
                    phase_steps, phase_fill, phase_font,
                    pos_labels, pos_axis, data, set_name):
    """1D sweep: rows = positions, columns = step names."""
    # Title row
    ws.cell(row=1, column=1,
            value="%s  |  %s %s  |  %s" % (set_name, label, stat_type, phase_name))
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(phase_steps) + 1)
    ws.cell(row=1, column=1).fill = phase_fill

    # Header: Position | Step1 | Step2 | ...
    ws.cell(row=2, column=1, value="Position_%s (mm)" % pos_axis)
    for si, step in enumerate(phase_steps):
        ws.cell(row=2, column=si + 2, value=step)
    _style_header_row(ws, 2, len(phase_steps) + 1, phase_fill, phase_font)

    # Data rows
    for ci, pl in enumerate(pos_labels):
        row_num = ci + 3
        _write_pos_label(ws, row_num, 1, pl)
        for si, step in enumerate(phase_steps):
            if pos_axis == "X":
                key = (step, ci, 0)
            else:
                key = (step, 0, ci)
            if key in data and label in data[key]:
                _write_value(ws, row_num, si + 2, data[key][label][stat_idx])


def _write_2d_sheet(ws, label, stat_type, stat_idx, phase_name,
                    phase_steps, phase_fill, phase_font,
                    x_labels, y_labels, data):
    """2D sweep: stacked Y x X blocks, one per step."""
    current_row = 1
    for step in phase_steps:
        # Step header
        ws.cell(row=current_row, column=1,
                value="%s  |  %s %s  |  %s" % (step, label, stat_type, phase_name))
        ws.cell(row=current_row, column=1).font = phase_font
        ws.cell(row=current_row, column=1).fill = phase_fill
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(x_labels) + 1)
        current_row += 1

        # Column headers (X positions)
        ws.cell(row=current_row, column=1, value="Y \\ X")
        for ci, xl in enumerate(x_labels):
            ws.cell(row=current_row, column=ci + 2, value=xl)
        _style_header_row(ws, current_row, len(x_labels) + 1)
        current_row += 1

        # Data rows (one per Y position)
        for yi_idx, yl in enumerate(y_labels):
            _write_pos_label(ws, current_row, 1, yl)
            for xi_idx in range(len(x_labels)):
                key = (step, xi_idx, yi_idx)
                if key in data and label in data[key]:
                    _write_value(ws, current_row, xi_idx + 2,
                                 data[key][label][stat_idx])
            current_row += 1

        current_row += 1  # blank row between step blocks


# ── Main ────────────────────────────────────────────────────────────────────
def convert_csv_to_excel(csv_path, xlsx_path=None):
    """Convert a consolidated CSV to a formatted, plot-ready Excel workbook."""
    base = os.path.splitext(os.path.basename(csv_path))[0]
    set_name = base.replace("_consolidated", "")

    if xlsx_path is None:
        xlsx_path = os.path.join(os.path.dirname(csv_path),
                                 "%s_Results.xlsx" % set_name)

    print("Reading: %s" % csv_path)
    meta, header, rows = parse_csv(csv_path)
    var_labels = detect_variable_labels(header)
    is_2d = meta.get("2D Sweep", "No").lower() == "yes"

    print("  Variables: %s" % ", ".join(var_labels))
    print("  2D Sweep:  %s" % is_2d)
    print("  Data rows: %d" % len(rows))

    x_labels, y_labels, step_order, data = build_data(rows, var_labels)
    print("  X positions: %d" % len(x_labels))
    print("  Y positions: %d" % len(y_labels))
    print("  Steps:       %d" % len(step_order))

    # Classify steps: Loading (odd index), Relaxation (even index)
    loading_steps = [s for i, s in enumerate(step_order) if i % 2 == 0]
    relaxation_steps = [s for i, s in enumerate(step_order) if i % 2 == 1]
    if len(step_order) == 1:
        loading_steps = step_order
        relaxation_steps = []

    print("  Loading steps:     %s" % ", ".join(loading_steps))
    print("  Relaxation steps:  %s" % ", ".join(relaxation_steps))

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    # Summary sheet
    write_summary(wb, set_name, var_labels, x_labels, y_labels, step_order,
                  loading_steps, relaxation_steps, data)

    # Per-variable sheets
    write_variable_sheets(wb, var_labels, x_labels, y_labels, step_order,
                          loading_steps, relaxation_steps, data, is_2d, set_name)

    wb.save(xlsx_path)

    # Count sheets
    sheet_names = wb.sheetnames
    print("\nCreated %d sheets:" % len(sheet_names))
    for name in sheet_names:
        print("  - %s" % name)
    print("\nSaved: %s" % xlsx_path)
    return xlsx_path


if __name__ == "__main__":
    if len(sys.argv) > 1:
        csv_file = sys.argv[1]
    else:
        candidates = [f for f in os.listdir(".")
                      if f.endswith("_consolidated.csv")]
        if not candidates:
            print("No *_consolidated.csv files found in current directory.")
            sys.exit(1)
        if len(candidates) > 1:
            print("Multiple CSV files found, processing all:")
            for c in candidates:
                convert_csv_to_excel(c)
            sys.exit(0)
        csv_file = candidates[0]

    convert_csv_to_excel(csv_file)
