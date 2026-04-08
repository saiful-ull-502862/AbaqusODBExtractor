"""
Flask backend for Abaqus ODB Field Output Extraction GUI.
Provides APIs for file browsing, ODB scanning, script generation, and execution.
"""

import json
import mimetypes
import os
import subprocess
import sys
import tempfile
import threading
import time
import queue
from pathlib import Path

from flask import Flask, request, jsonify, Response, send_from_directory
from flask_cors import CORS

# Fix MIME types on Windows (Python may serve .js as text/plain)
mimetypes.add_type("application/javascript", ".js")
mimetypes.add_type("text/css", ".css")

app = Flask(__name__, static_folder="static", static_url_path="")
CORS(app)

# Global state for running process
_process = None
_process_lock = threading.Lock()
_log_queue = queue.Queue()
_run_start_time = None
_run_status = "idle"  # idle, running, completed, failed, cancelled


# ============================================================================
# FILE BROWSING (uses tkinter dialogs)
# ============================================================================

@app.route("/api/browse-file", methods=["POST"])
def browse_file():
    """Open a native file dialog to select an ODB file."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        file_path = filedialog.askopenfilename(
            title="Select Abaqus ODB File",
            filetypes=[("Abaqus ODB files", "*.odb"), ("All files", "*.*")],
        )
        root.destroy()
        return jsonify({"path": file_path if file_path else ""})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/browse-dir", methods=["POST"])
def browse_dir():
    """Open a native directory dialog."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        initial_dir = request.json.get("initialDir", "")
        dir_path = filedialog.askdirectory(
            title="Select Output Directory",
            initialdir=initial_dir if initial_dir else None,
        )
        root.destroy()
        return jsonify({"path": dir_path if dir_path else ""})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================================
# ABAQUS DETECTION
# ============================================================================

ABAQUS_SEARCH_PATHS = [
    "abaqus",
    # Common SIMULIA installations (various versions)
    r"C:\SIMULIA\Abaqus\Commands\abaqus.bat",
    r"C:\SIMULIA\Commands\abaqus.bat",
    r"C:\SIMULIA\EstProducts\Commands\abaqus.bat",
    # Abaqus 2020-2025 (Dassault Systemes newer layout)
    r"C:\Program Files\Dassault Systemes\SimulationServices\V6R2023\win_b64\code\bin\ABQLauncher.exe",
    r"C:\Program Files\Dassault Systemes\SimulationServices\V6R2024\win_b64\code\bin\ABQLauncher.exe",
    r"C:\Program Files\Dassault Systemes\SimulationServices\V6R2025\win_b64\code\bin\ABQLauncher.exe",
    r"C:\Program Files\Dassault Systemes\SimulationServices\V6R2022\win_b64\code\bin\ABQLauncher.exe",
    r"C:\Program Files\Dassault Systemes\SimulationServices\V6R2021\win_b64\code\bin\ABQLauncher.exe",
    r"C:\Program Files\Dassault Systemes\SimulationServices\V6R2020\win_b64\code\bin\ABQLauncher.exe",
    r"C:\Program Files\Dassault Systemes\SimulationServices\V6R2014\win_b64\code\bin\ABQLauncher.exe",
]


@app.route("/api/detect-abaqus", methods=["POST"])
def detect_abaqus():
    """Try to auto-detect Abaqus installation."""
    custom_path = request.json.get("customPath", "")
    paths_to_try = [custom_path] if custom_path else ABAQUS_SEARCH_PATHS

    for abaqus_cmd in paths_to_try:
        if not abaqus_cmd:
            continue
        try:
            use_shell = sys.platform == "win32" and abaqus_cmd.lower().endswith(".bat")
            cmd = [abaqus_cmd, "information=release"]
            if use_shell:
                cmd = subprocess.list2cmdline(cmd)
            result = subprocess.run(
                cmd,
                capture_output=True, text=True, timeout=15,
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
                shell=use_shell,
            )
            output = result.stdout + result.stderr
            if "Abaqus" in output or "abaqus" in output.lower():
                version_line = ""
                for line in output.splitlines():
                    if "abaqus" in line.lower() and any(c.isdigit() for c in line):
                        version_line = line.strip()
                        break
                return jsonify({
                    "found": True,
                    "path": abaqus_cmd,
                    "version": version_line or "Abaqus detected",
                    "output": output.strip()[:500],
                })
        except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
            continue

    return jsonify({"found": False, "error": "Abaqus not found. Please specify the path manually."})


# ============================================================================
# ODB PROBE / SCAN
# ============================================================================

PROBE_SCRIPT_TEMPLATE = r'''# Auto-generated ODB probe script for Abaqus Python 2.7
# Usage: abaqus cae noGUI=probe_odb.py
import json
import os
import sys

from odbAccess import *

ODB_PATH = r'{odb_path}'
OUTPUT_JSON = r'{output_json}'

print('Opening ODB for scanning: %s' % ODB_PATH)
odb = openOdb(path=ODB_PATH, readOnly=True)

result = {{
    'instances': [],
    'elementSets': [],
    'steps': [],
    'fieldOutputs': [],
    'boundingBox': None,
}}

# Instance names
for inst_name in odb.rootAssembly.instances.keys():
    result['instances'].append(inst_name)

# Element sets - rootAssembly level
ra_keys = list(odb.rootAssembly.elementSets.keys())
for k in ra_keys:
    elset = odb.rootAssembly.elementSets[k]
    count = 0
    try:
        first = elset.elements[0]
        if hasattr(first, 'label'):
            count = len(elset.elements)
        else:
            for arr in elset.elements:
                count += len(arr)
    except Exception:
        pass
    result['elementSets'].append({{
        'name': k,
        'source': 'rootAssembly',
        'elementCount': count,
    }})

# Element sets - instance level
for inst_name in odb.rootAssembly.instances.keys():
    inst = odb.rootAssembly.instances[inst_name]
    inst_keys = list(inst.elementSets.keys())
    for k in inst_keys:
        elset = inst.elementSets[k]
        count = 0
        try:
            first = elset.elements[0]
            if hasattr(first, 'label'):
                count = len(elset.elements)
            else:
                for arr in elset.elements:
                    count += len(arr)
        except Exception:
            pass
        result['elementSets'].append({{
            'name': '%s.%s' % (inst_name, k),
            'source': 'instance',
            'elementCount': count,
        }})

# Steps and frames
for step_name in odb.steps.keys():
    step = odb.steps[step_name]
    result['steps'].append({{
        'name': step_name,
        'frameCount': len(step.frames),
        'totalTime': step.totalTime,
    }})

# Field outputs (from the first step's last frame)
step_keys = list(odb.steps.keys())
if step_keys:
    first_step = odb.steps[step_keys[0]]
    if len(first_step.frames) > 0:
        frame = first_step.frames[-1]
        for fo_key in frame.fieldOutputs.keys():
            fo = frame.fieldOutputs[fo_key]
            comps = list(fo.componentLabels) if fo.componentLabels else []
            # Check available invariants
            invariants = []
            inv_names = ['mises', 'tresca', 'press', 'inv3',
                         'maxPrincipal', 'midPrincipal', 'minPrincipal',
                         'maxInPlanePrincipal', 'minInPlanePrincipal',
                         'outOfPlanePrincipal']
            if len(fo.values) > 0:
                fv0 = fo.values[0]
                for inv_attr in inv_names:
                    try:
                        val = getattr(fv0, inv_attr, None)
                        if val is not None:
                            invariants.append(inv_attr)
                    except Exception:
                        pass

            # Detect if nodal
            is_nodal = False
            if len(fo.values) > 0:
                fv0 = fo.values[0]
                if (hasattr(fv0, 'nodeLabel') and fv0.nodeLabel is not None and
                    (not hasattr(fv0, 'elementLabel') or fv0.elementLabel is None or
                     fv0.elementLabel == 0)):
                    is_nodal = True

            result['fieldOutputs'].append({{
                'key': fo_key,
                'description': fo.description if hasattr(fo, 'description') else '',
                'components': comps,
                'invariants': invariants,
                'isNodal': is_nodal,
            }})

# Bounding box from first instance nodes
if result['instances']:
    inst = odb.rootAssembly.instances[result['instances'][0]]
    x_min, x_max = 1e30, -1e30
    y_min, y_max = 1e30, -1e30
    z_min, z_max = 1e30, -1e30
    for node in inst.nodes:
        c = node.coordinates
        if c[0] < x_min: x_min = c[0]
        if c[0] > x_max: x_max = c[0]
        if c[1] < y_min: y_min = c[1]
        if c[1] > y_max: y_max = c[1]
        if len(c) > 2:
            if c[2] < z_min: z_min = c[2]
            if c[2] > z_max: z_max = c[2]
    result['boundingBox'] = {{
        'xMin': x_min, 'xMax': x_max,
        'yMin': y_min, 'yMax': y_max,
        'zMin': z_min if z_min < 1e29 else 0,
        'zMax': z_max if z_max > -1e29 else 0,
    }}

odb.close()
print('ODB scan complete.')

# Write JSON output
with open(OUTPUT_JSON, 'w') as f:
    json.dump(result, f, indent=2)

print('Results written to: %s' % OUTPUT_JSON)
'''


@app.route("/api/scan-odb", methods=["POST"])
def scan_odb():
    """Generate and run the ODB probe script, return metadata."""
    data = request.json
    odb_path = data.get("odbPath", "")
    abaqus_cmd = data.get("abaqusPath", "abaqus")

    if not odb_path or not os.path.isfile(odb_path):
        return jsonify({"error": "ODB file not found: %s" % odb_path}), 400

    # Create temp files for probe script and JSON output
    temp_dir = tempfile.mkdtemp(prefix="abq_probe_")
    probe_script = os.path.join(temp_dir, "probe_odb.py")
    output_json = os.path.join(temp_dir, "odb_metadata.json")

    # Generate probe script
    script_content = PROBE_SCRIPT_TEMPLATE.format(
        odb_path=odb_path.replace("\\", "\\\\"),
        output_json=output_json.replace("\\", "\\\\"),
    )
    with open(probe_script, "w") as f:
        f.write(script_content)

    # Run probe script via Abaqus
    try:
        flags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
        use_shell = sys.platform == "win32" and abaqus_cmd.lower().endswith(".bat")
        cmd = [abaqus_cmd, "cae", "noGUI=%s" % probe_script]
        if use_shell:
            cmd = subprocess.list2cmdline(cmd)
        result = subprocess.run(
            cmd,
            capture_output=True, text=True, timeout=300,
            cwd=temp_dir, creationflags=flags, shell=use_shell,
        )

        if not os.path.isfile(output_json):
            return jsonify({
                "error": "Probe script did not produce output. Abaqus output:\n%s\n%s"
                         % (result.stdout[-2000:], result.stderr[-2000:])
            }), 500

        with open(output_json, "r") as f:
            metadata = json.load(f)

        return jsonify(metadata)

    except subprocess.TimeoutExpired:
        return jsonify({"error": "ODB scan timed out after 5 minutes."}), 500
    except FileNotFoundError:
        return jsonify({"error": "Abaqus command not found: %s" % abaqus_cmd}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        # Clean up temp files
        try:
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception:
            pass


# ============================================================================
# SCRIPT GENERATION
# ============================================================================

EXTRACTION_SCRIPT_TEMPLATE = r'''# ============================================================================
# Abaqus Python Script: Extract Field Outputs from Multiple Regions
# Auto-generated by Abaqus ODB Extractor
# Usage: abaqus cae noGUI=extract_region_sweep.py
#        or: abaqus python extract_region_sweep.py
#        or inside CAE: File > Run Script
# Compatible with Abaqus Python 2.7 (Abaqus 6.14+)
# ============================================================================

from odbAccess import *
from abaqusConstants import *
import os
import sys
import time
import json

# ============================================================================
# USER CONFIGURATION (Auto-Generated)
# ============================================================================

ODB_PATH = r'{odb_path}'
OUTPUT_DIR = r'{output_dir}'
INSTANCE_NAME = '{instance_name}'

MODEL_WIDTH_X  = {model_width}    # mm
MODEL_HEIGHT_Y = {model_height}    # mm

ELEMENT_SET_NAMES = [
{element_sets}
]

REGIONS = [
{regions}
]

X_LABELS = {x_labels}
Y_LABELS = {y_labels}
IS_2D_SWEEP = {is_2d}

OUTPUTS_TO_EXTRACT = [
{outputs}
]

STEPS_TO_EXTRACT = {steps_config}
FRAME_SELECTION  = '{frame_selection}'

# ============================================================================
# INVARIANT AND COMPONENT HELPERS
# ============================================================================

INVARIANT_MAP = {{
    'Mises':                    MISES,
    'Tresca':                   TRESCA,
    'Press':                    PRESS,
    'INV3':                     INV3,
    'Max. Principal':           MAX_PRINCIPAL,
    'Mid. Principal':           MID_PRINCIPAL,
    'Min. Principal':           MIN_PRINCIPAL,
    'Max. In-Plane Principal':  MAX_INPLANE_PRINCIPAL,
    'Min. In-Plane Principal':  MIN_INPLANE_PRINCIPAL,
    'Out-of-Plane Principal':   OUTOFPLANE_PRINCIPAL,
    'Magnitude':                MAGNITUDE,
}}

INVARIANT_ATTR = {{
    'Mises':                    'mises',
    'Tresca':                   'tresca',
    'Press':                    'press',
    'INV3':                     'inv3',
    'Max. Principal':           'maxPrincipal',
    'Mid. Principal':           'midPrincipal',
    'Min. Principal':           'minPrincipal',
    'Max. In-Plane Principal':  'maxInPlanePrincipal',
    'Min. In-Plane Principal':  'minInPlanePrincipal',
    'Out-of-Plane Principal':   'outOfPlanePrincipal',
    'Magnitude':                'magnitude',
}}


def is_invariant(comp_or_inv):
    return comp_or_inv in INVARIANT_MAP


def extract_value(fv, comp_or_inv, field_output):
    """Extract scalar value from a FieldValue."""
    if comp_or_inv is None:
        return fv.data

    if is_invariant(comp_or_inv):
        attr_name = INVARIANT_ATTR[comp_or_inv]
        return getattr(fv, attr_name)

    comp_labels = field_output.componentLabels
    if comp_or_inv in comp_labels:
        idx = comp_labels.index(comp_or_inv)
        return fv.data[idx]
    return None


# ============================================================================
# MAIN
# ============================================================================

print('=' * 70)
print('  ABAQUS FIELD OUTPUT EXTRACTION - REGION SWEEP')
print('  Auto-generated by Abaqus ODB Extractor')
print('=' * 70)

start_time = time.time()

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)
    print('\nCreated output directory: %s' % OUTPUT_DIR)

print('\nOpening ODB: %s' % ODB_PATH)
odb = openOdb(path=ODB_PATH, readOnly=True)

instance = odb.rootAssembly.instances[INSTANCE_NAME.upper()]

# Build node coordinate map
print('\nBuilding node coordinate map...')
node_coord_map = {{}}
for node in instance.nodes:
    node_coord_map[node.label] = node.coordinates
print('  Nodes indexed: %d' % len(node_coord_map))

# Cache element set keys (Abaqus 6.14 safe)
ra_elset_keys = list(odb.rootAssembly.elementSets.keys())
inst_elset_keys = list(instance.elementSets.keys())

ra_elset_dict = {{}}
for k in ra_elset_keys:
    ra_elset_dict[k] = k
    ra_elset_dict[k.upper()] = k

inst_elset_dict = {{}}
for k in inst_elset_keys:
    inst_elset_dict[k] = k
    inst_elset_dict[k.upper()] = k

# Determine steps
if STEPS_TO_EXTRACT == 'ALL':
    step_names = list(odb.steps.keys())
else:
    step_names = STEPS_TO_EXTRACT

total_regions = len(REGIONS)
total_elsets = len(ELEMENT_SET_NAMES)

print('\nSteps to process (%d): %s' % (len(step_names), ', '.join(step_names)))
print('Regions to process: %d' % total_regions)
print('Element sets to process: %d' % total_elsets)
print('Output variables: %d' % len(OUTPUTS_TO_EXTRACT))
print('2D sweep: %s' % ('Yes' if IS_2D_SWEEP else 'No'))

header_labels = [out[2] for out in OUTPUTS_TO_EXTRACT]

# ============================================================================
# PRE-COMPUTE: element centroids and connectivity (ONE pass over all elements)
# ============================================================================
print('\nPre-computing element centroids (single pass)...')
precomp_start = time.time()
elem_centroid = {{}}   # label -> (cx, cy)
elem_node_map = {{}}   # label -> list of node labels
for elem in instance.elements:
    connectivity = elem.connectivity
    n_nodes = len(connectivity)
    cx, cy = 0.0, 0.0
    for nLabel in connectivity:
        coords = node_coord_map[nLabel]
        cx += coords[0]
        cy += coords[1]
    cx /= n_nodes
    cy /= n_nodes
    elem_centroid[elem.label] = (cx, cy)
    elem_node_map[elem.label] = list(connectivity)
print('  Elements pre-computed: %d (%.1f sec)' % (len(elem_centroid), time.time() - precomp_start))

sys.stdout.flush()


def find_element_set(set_name):
    """Find element set by name with multiple lookup strategies."""
    if '.' in set_name:
        inst_part, set_part = set_name.rsplit('.', 1)
    else:
        inst_part = INSTANCE_NAME
        set_part = set_name

    candidates_ra = [set_name, set_name.upper(),
                     '%s.%s' % (inst_part.upper(), set_part.upper()),
                     set_part, set_part.upper()]
    candidates_inst = [set_part, set_part.upper(), set_name, set_name.upper()]

    for candidate in candidates_ra:
        if candidate in ra_elset_dict:
            actual_key = ra_elset_dict[candidate]
            return odb.rootAssembly.elementSets[actual_key], actual_key

    for candidate in candidates_inst:
        if candidate in inst_elset_dict:
            actual_key = inst_elset_dict[candidate]
            return instance.elementSets[actual_key], actual_key

    # Partial match
    set_part_upper = set_part.upper()
    for k in ra_elset_keys:
        if k.upper().endswith(set_part_upper):
            return odb.rootAssembly.elementSets[k], k
    for k in inst_elset_keys:
        if k.upper().endswith(set_part_upper):
            return instance.elementSets[k], k

    return None, None


def get_element_labels(elset):
    """Collect element labels from an element set."""
    labels = set()
    try:
        first_item = elset.elements[0]
        if hasattr(first_item, 'label'):
            for elem in elset.elements:
                labels.add(elem.label)
        elif hasattr(first_item, '__iter__') or hasattr(first_item, '__getitem__'):
            for elemArray in elset.elements:
                for elem in elemArray:
                    labels.add(elem.label)
        else:
            for elem in elset.elements:
                labels.add(elem.label)
    except (TypeError, IndexError):
        for item in elset.elements:
            if hasattr(item, 'label'):
                labels.add(item.label)
            else:
                for elem in item:
                    labels.add(elem.label)
    return labels


# ============================================================================
# PROCESS EACH ELEMENT SET
# ============================================================================

total_csv_files = 0

for elset_name in ELEMENT_SET_NAMES:
    print('\n' + '=' * 70)
    print('Processing Element Set: %s' % elset_name)
    print('=' * 70)

    elset, found_key = find_element_set(elset_name)
    if elset is None:
        print('*** ERROR: Element set "%s" not found! Skipping. ***' % elset_name)
        continue

    print('  Found as: "%s"' % found_key)
    elset_element_labels = get_element_labels(elset)
    print('  Total elements in set: %d' % len(elset_element_labels))

    # -----------------------------------------------------------------
    # PRE-ASSIGN elements to regions (single pass over elset elements)
    # -----------------------------------------------------------------
    print('  Assigning elements to %d regions...' % total_regions)
    assign_start = time.time()

    # region_idx -> set of element labels
    region_elem_map = {{}}
    # region_idx -> set of node labels
    region_node_map = {{}}
    for ri in range(total_regions):
        region_elem_map[ri] = set()
        region_node_map[ri] = set()

    for elabel in elset_element_labels:
        if elabel not in elem_centroid:
            continue
        cx, cy = elem_centroid[elabel]
        nodes = elem_node_map[elabel]
        for ri, region_info in enumerate(REGIONS):
            if (region_info['x_min'] <= cx <= region_info['x_max'] and
                region_info['y_min'] <= cy <= region_info['y_max']):
                region_elem_map[ri].add(elabel)
                for nLabel in nodes:
                    region_node_map[ri].add(nLabel)

    assign_time = time.time() - assign_start
    non_empty = sum([1 for ri in range(total_regions) if len(region_elem_map[ri]) > 0])
    print('  Region assignment done: %d/%d non-empty (%.1f sec)' % (non_empty, total_regions, assign_time))
    sys.stdout.flush()

    # -----------------------------------------------------------------
    # EXTRACT: iterate steps/frames ONCE, gather data for ALL regions
    # -----------------------------------------------------------------
    # Instead of looping regions->steps->frames->fieldValues, we loop
    # steps->frames->fieldValues ONCE and bucket into all matching regions.
    # -----------------------------------------------------------------
    print('  Extracting field outputs across all regions...')
    extract_start = time.time()

    # Build element-to-regions and node-to-regions lookup
    elem_to_regions = {{}}
    node_to_regions = {{}}
    for ri in range(total_regions):
        for el in region_elem_map[ri]:
            elem_to_regions.setdefault(el, []).append(ri)
        for nl in region_node_map[ri]:
            node_to_regions.setdefault(nl, []).append(ri)

    # Storage: (region_idx, step_name, frame_idx) -> {{label: (max_val, min_val)}}
    results = {{}}
    frame_meta = {{}}  # (step_name, frame_idx) -> (frame_time, total_time)

    cumulative_time = 0.0
    for si, step_name in enumerate(step_names):
        step = odb.steps[step_name]

        if FRAME_SELECTION == 'LAST':
            frames_to_process = [step.frames[-1]]
        else:
            frames_to_process = list(step.frames)

        print('    Step %d/%d: %-12s | Frames: %d' % (si + 1, len(step_names), step_name, len(frames_to_process)))
        sys.stdout.flush()

        for frame in frames_to_process:
            frame_idx  = frame.frameId
            frame_time = frame.frameValue
            total_time = cumulative_time + frame_time
            frame_meta[(step_name, frame_idx)] = (frame_time, total_time)

            # Initialize results for all non-empty regions
            for ri in range(total_regions):
                if len(region_elem_map[ri]) == 0:
                    continue
                key = (ri, step_name, frame_idx)
                results[key] = {{}}
                for label in header_labels:
                    results[key][label] = (None, None)  # (max, min)

            for (field_key, comp_or_inv, label) in OUTPUTS_TO_EXTRACT:
                if field_key not in frame.fieldOutputs:
                    continue

                field_output = frame.fieldOutputs[field_key]

                is_nodal_field = False
                if len(field_output.values) > 0:
                    fv0 = field_output.values[0]
                    if (hasattr(fv0, 'nodeLabel') and fv0.nodeLabel is not None and
                        (not hasattr(fv0, 'elementLabel') or fv0.elementLabel is None or
                         fv0.elementLabel == 0)):
                        is_nodal_field = True

                # Single pass over field values, bucket into all matching regions
                for fv in field_output.values:
                    if is_nodal_field:
                        nl = fv.nodeLabel
                        if nl not in node_to_regions:
                            continue
                        matched_regions = node_to_regions[nl]
                    else:
                        if not hasattr(fv, 'elementLabel') or fv.elementLabel is None:
                            continue
                        el = fv.elementLabel
                        if el not in elem_to_regions:
                            continue
                        matched_regions = elem_to_regions[el]

                    try:
                        val = extract_value(fv, comp_or_inv, field_output)
                    except Exception:
                        continue
                    if val is None:
                        continue

                    for ri in matched_regions:
                        key = (ri, step_name, frame_idx)
                        if key not in results:
                            continue
                        cur_max, cur_min = results[key][label]
                        if cur_max is None or val > cur_max:
                            cur_max = val
                        if cur_min is None or val < cur_min:
                            cur_min = val
                        results[key][label] = (cur_max, cur_min)

        cumulative_time += step.totalTime

    extract_time = time.time() - extract_start
    print('  Extraction done (%.1f sec)' % extract_time)
    sys.stdout.flush()

    # -----------------------------------------------------------------
    # BUILD CSV ROWS from results dict
    # -----------------------------------------------------------------
    csv_rows = []
    for region_idx, region_info in enumerate(REGIONS):
        if len(region_elem_map[region_idx]) == 0:
            continue
        xi = region_info.get('xi', 0)
        yi = region_info.get('yi', 0)
        x_label = region_info.get('x_label', '')
        y_label = region_info.get('y_label', '')

        cumulative_time = 0.0
        for step_name in step_names:
            step = odb.steps[step_name]
            if FRAME_SELECTION == 'LAST':
                frames_to_process = [step.frames[-1]]
            else:
                frames_to_process = list(step.frames)

            for frame in frames_to_process:
                frame_idx = frame.frameId
                key = (region_idx, step_name, frame_idx)
                if key not in results:
                    continue
                fm = frame_meta.get((step_name, frame_idx), (0.0, 0.0))
                row = [step_name, frame_idx, '%.6e' % fm[0], '%.6e' % fm[1],
                       region_idx, xi, yi, x_label, y_label]
                for label in header_labels:
                    mv, nv = results[key].get(label, (None, None))
                    row.append('%.8e' % mv if mv is not None else 'N/A')
                    row.append('%.8e' % nv if nv is not None else 'N/A')
                csv_rows.append(row)

    print('  [%d/%d] Regions processed, %d data rows' % (non_empty, total_regions, len(csv_rows)))
    sys.stdout.flush()

    # Write consolidated CSV
    set_short = elset_name.split('.')[-1] if '.' in elset_name else elset_name
    csv_filename = '%s_consolidated.csv' % set_short
    csv_path = os.path.join(OUTPUT_DIR, csv_filename)

    csv_header = ['Step', 'Frame', 'StepTime_s', 'TotalTime_s',
                  'RegionIdx', 'Xi', 'Yi', 'XLabel', 'YLabel']
    for label in header_labels:
        csv_header.append('%s_Max' % label)
        csv_header.append('%s_Min' % label)

    with open(csv_path, 'w') as f:
        f.write('# Abaqus Field Output Extraction - Consolidated\n')
        f.write('# ODB: %s\n' % ODB_PATH)
        f.write('# Element Set: %s\n' % elset_name)
        f.write('# Total Regions: %d\n' % total_regions)
        f.write('# 2D Sweep: %s\n' % ('Yes' if IS_2D_SWEEP else 'No'))
        f.write('# X Labels: %s\n' % '; '.join([str(x) for x in X_LABELS]))
        f.write('# Y Labels: %s\n' % '; '.join([str(y) for y in Y_LABELS]))
        f.write('#\n')
        f.write(','.join(csv_header) + '\n')
        for row in csv_rows:
            f.write(','.join([str(v) for v in row]) + '\n')

    print('\n  >> Wrote consolidated CSV: %s (%d data rows)' % (csv_filename, len(csv_rows)))
    total_csv_files += 1

# ============================================================================
# SUMMARY
# ============================================================================
elapsed = time.time() - start_time
print('\n' + '=' * 70)
print('  EXTRACTION COMPLETE')
print('=' * 70)
print('  Consolidated CSV files written: %d' % total_csv_files)
print('  Output directory: %s' % OUTPUT_DIR)
print('  Elapsed time: %.1f seconds' % elapsed)
print('=' * 70)

# Write metadata JSON for post-processing
metadata = {{
    'output_dir': OUTPUT_DIR,
    'element_sets': ELEMENT_SET_NAMES,
    'header_labels': header_labels,
    'x_labels': X_LABELS,
    'y_labels': Y_LABELS,
    'is_2d': IS_2D_SWEEP,
    'step_names': step_names,
}}
meta_path = os.path.join(OUTPUT_DIR, '_extraction_metadata.json')
with open(meta_path, 'w') as f:
    json.dump(metadata, f, indent=2)
print('Metadata written to: %s' % meta_path)

odb.close()
print('\nDone. ODB closed.')
# ============================================================================
# AUTO-GENERATE EXCEL WORKBOOK (runs via system Python 3, not Abaqus Python)
# ============================================================================
import subprocess as _sp
import base64 as _b64

# Embedded format_to_excel.py (base64-encoded)
_CONVERTER_B64 = (
    "IiIiCkNvbnZlcnQgY29uc29saWRhdGVkIENTViBmcm9tIEFiYXF1cyBleHRyYWN0aW9uIGludG8gcGxvdC1yZWFkeSBFeGNlbCB3b3JrYm9va3MuCgpTdHJ1Y3R1cmU6CiAgLSBTdW1tYXJ5IHNoZWV0OiAgcGVhayB2YWx1ZXMgcGVyIHN0ZXAgcGVyIHZhcmlhYmxlCiAgLSBQZXIgdmFyaWFibGU6ICAgc2VwYXJhdGUgc2hlZXRzIGZvciBMb2FkaW5nICYgUmVsYXhhdGlvbiBwaGFzZXMKICAgICAgICAgICAgICAgICAgICAxRCBzd2VlcCA9PiByb3dzID0gWCBwb3NpdGlvbnMsIGNvbHVtbnMgPSBzdGVwIG5hbWVzCiAgICAgICAgICAgICAgICAgICAgMkQgc3dlZXAgPT4gc3RhY2tlZCBZIHggWCBibG9ja3MgcGVyIHN0ZXAKICAtIEFsbFN0ZXBzIHNoZWV0czogYWxsIHN0ZXBzIGluIG9uZSBzaGVldCAoTG9hZGluZyBjb2xvci1jb2RlZCB2cyBSZWxheGF0aW9uKQoKU3RlcHMgYXJlIGNsYXNzaWZpZWQgYXV0b21hdGljYWxseToKICAtIExvYWRpbmc6ICAgICBvZGQgc3RlcHMgKDFzdCwgM3JkLCA1dGgsIC4uLikKICAtIFJlbGF4YXRpb246ICBldmVuIHN0ZXBzICgybmQsIDR0aCwgNnRoLCAuLi4pCgpVc2FnZToKICAgIHB5dGhvbiBmb3JtYXRfdG9fZXhjZWwucHkgICAgICAgICAgICAgICAgICAgICAgICAgICAgIyBhdXRvLWZpbmRzICpfY29uc29saWRhdGVkLmNzdgogICAgcHl0aG9uIGZvcm1hdF90b19leGNlbC5weSAgRUNNX2NvbnNvbGlkYXRlZC5jc3YgICAgICAjIHNwZWNpZmljIGZpbGUKIiIiCgppbXBvcnQgb3MKaW1wb3J0IHN5cwppbXBvcnQgY3N2CmZyb20gY29sbGVjdGlvbnMgaW1wb3J0IE9yZGVyZWREaWN0Cgpmcm9tIG9wZW5weXhsIGltcG9ydCBXb3JrYm9vawpmcm9tIG9wZW5weXhsLnN0eWxlcyBpbXBvcnQgRm9udCwgQWxpZ25tZW50LCBQYXR0ZXJuRmlsbCwgQm9yZGVyLCBTaWRlCmZyb20gb3BlbnB5eGwudXRpbHMgaW1wb3J0IGdldF9jb2x1bW5fbGV0dGVyCgoKIyDilIDilIAgU3R5bGUgZGVmaW5pdGlvbnMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACkhFQURFUl9GSUxMX0RBUksgPSBQYXR0ZXJuRmlsbChzdGFydF9jb2xvcj0iMkQzNzQ4IiwgZW5kX2NvbG9yPSIyRDM3NDgiLCBmaWxsX3R5cGU9InNvbGlkIikKSEVBREVSX0ZPTlRfV0hJVEUgPSBGb250KGJvbGQ9VHJ1ZSwgc2l6ZT0xMCwgY29sb3I9IkZGRkZGRiIpCkxPQURJTkdfRklMTCA9IFBhdHRlcm5GaWxsKHN0YXJ0X2NvbG9yPSIxQTM2NUQiLCBlbmRfY29sb3I9IjFBMzY1RCIsIGZpbGxfdHlwZT0ic29saWQiKQpMT0FESU5HX0ZPTlQgPSBGb250KGJvbGQ9VHJ1ZSwgc2l6ZT0xMCwgY29sb3I9IjkwQ0RGNCIpClJFTEFYX0ZJTEwgPSBQYXR0ZXJuRmlsbChzdGFydF9jb2xvcj0iMzIyNjU5IiwgZW5kX2NvbG9yPSIzMjI2NTkiLCBmaWxsX3R5cGU9InNvbGlkIikKUkVMQVhfRk9OVCA9IEZvbnQoYm9sZD1UcnVlLCBzaXplPTEwLCBjb2xvcj0iRDZCQ0ZBIikKVEhJTl9CT1JERVIgPSBCb3JkZXIoCiAgICBsZWZ0PVNpZGUoc3R5bGU9InRoaW4iLCBjb2xvcj0iQ0NDQ0NDIiksCiAgICByaWdodD1TaWRlKHN0eWxlPSJ0aGluIiwgY29sb3I9IkNDQ0NDQyIpLAogICAgdG9wPVNpZGUoc3R5bGU9InRoaW4iLCBjb2xvcj0iQ0NDQ0NDIiksCiAgICBib3R0b209U2lkZShzdHlsZT0idGhpbiIsIGNvbG9yPSJDQ0NDQ0MiKSwKKQpQT1NfTEFCRUxfRk9OVCA9IEZvbnQoYm9sZD1GYWxzZSwgc2l6ZT0xMCwgY29sb3I9IkEwQUVDMCIpClBPU19MQUJFTF9GSUxMID0gUGF0dGVybkZpbGwoc3RhcnRfY29sb3I9IjFBMjAyQyIsIGVuZF9jb2xvcj0iMUEyMDJDIiwgZmlsbF90eXBlPSJzb2xpZCIpClNVTU1BUllfSEVBREVSX0ZJTEwgPSBQYXR0ZXJuRmlsbChzdGFydF9jb2xvcj0iMUEyMDJDIiwgZW5kX2NvbG9yPSIxQTIwMkMiLCBmaWxsX3R5cGU9InNvbGlkIikKU1VNTUFSWV9IRUFERVJfRk9OVCA9IEZvbnQoYm9sZD1UcnVlLCBzaXplPTExLCBjb2xvcj0iRkZGRkZGIikKCgpkZWYgX3N0eWxlX2hlYWRlcl9yb3cod3MsIHJvd19udW0sIG51bV9jb2xzLCBmaWxsPU5vbmUsIGZvbnQ9Tm9uZSk6CiAgICBmb3IgYyBpbiByYW5nZSgxLCBudW1fY29scyArIDEpOgogICAgICAgIGNlbGwgPSB3cy5jZWxsKHJvdz1yb3dfbnVtLCBjb2x1bW49YykKICAgICAgICBjZWxsLmZvbnQgPSBmb250IG9yIEhFQURFUl9GT05UX1dISVRFCiAgICAgICAgY2VsbC5maWxsID0gZmlsbCBvciBIRUFERVJfRklMTF9EQVJLCiAgICAgICAgY2VsbC5ib3JkZXIgPSBUSElOX0JPUkRFUgogICAgICAgIGNlbGwuYWxpZ25tZW50ID0gQWxpZ25tZW50KGhvcml6b250YWw9ImNlbnRlciIpCgoKZGVmIF93cml0ZV92YWx1ZSh3cywgcm93X251bSwgY29sX251bSwgdmFsX3N0cik6CiAgICB0cnk6CiAgICAgICAgdmFsID0gZmxvYXQodmFsX3N0cikKICAgICAgICBjZWxsID0gd3MuY2VsbChyb3c9cm93X251bSwgY29sdW1uPWNvbF9udW0sIHZhbHVlPXZhbCkKICAgICAgICBjZWxsLm51bWJlcl9mb3JtYXQgPSAnMC4wMDAwRSswMCcKICAgIGV4Y2VwdCAoVmFsdWVFcnJvciwgVHlwZUVycm9yKToKICAgICAgICBjZWxsID0gd3MuY2VsbChyb3c9cm93X251bSwgY29sdW1uPWNvbF9udW0sIHZhbHVlPXZhbF9zdHIpCiAgICBjZWxsLmJvcmRlciA9IFRISU5fQk9SREVSCiAgICBjZWxsLmFsaWdubWVudCA9IEFsaWdubWVudChob3Jpem9udGFsPSJjZW50ZXIiKQoKCmRlZiBfd3JpdGVfcG9zX2xhYmVsKHdzLCByb3dfbnVtLCBjb2xfbnVtLCB2YWx1ZSk6CiAgICBjZWxsID0gd3MuY2VsbChyb3c9cm93X251bSwgY29sdW1uPWNvbF9udW0sIHZhbHVlPXZhbHVlKQogICAgY2VsbC5mb250ID0gUE9TX0xBQkVMX0ZPTlQKICAgIGNlbGwuZmlsbCA9IFBPU19MQUJFTF9GSUxMCiAgICBjZWxsLmJvcmRlciA9IFRISU5fQk9SREVSCiAgICBjZWxsLmFsaWdubWVudCA9IEFsaWdubWVudChob3Jpem9udGFsPSJjZW50ZXIiKQoKCmRlZiBfYXV0b19maXQod3MpOgogICAgZm9yIGNvbCBpbiB3cy5jb2x1bW5zOgogICAgICAgIG1heF9sZW4gPSAwCiAgICAgICAgY29sX2xldHRlciA9IE5vbmUKICAgICAgICBmb3IgY2VsbCBpbiBjb2w6CiAgICAgICAgICAgIGlmIGhhc2F0dHIoY2VsbCwgJ2NvbHVtbl9sZXR0ZXInKToKICAgICAgICAgICAgICAgIGNvbF9sZXR0ZXIgPSBjZWxsLmNvbHVtbl9sZXR0ZXIKICAgICAgICAgICAgaWYgY2VsbC52YWx1ZSBpcyBub3QgTm9uZSBhbmQgbm90IGlzaW5zdGFuY2UoY2VsbCwgdHlwZShOb25lKSk6CiAgICAgICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICAgICAgbWF4X2xlbiA9IG1heChtYXhfbGVuLCBsZW4oc3RyKGNlbGwudmFsdWUpKSkKICAgICAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAgICAgICAgICAgICAgICAgcGFzcwogICAgICAgIGlmIGNvbF9sZXR0ZXI6CiAgICAgICAgICAgIHdzLmNvbHVtbl9kaW1lbnNpb25zW2NvbF9sZXR0ZXJdLndpZHRoID0gbWluKG1heF9sZW4gKyAzLCAyMikKCgpkZWYgX3NhZmVfc2hlZXRfbmFtZShuYW1lLCBtYXhfbGVuPTMxKToKICAgIGZvciBjaCBpbiBbJ1xcJywgJy8nLCAnKicsICc/JywgJzonLCAnWycsICddJ106CiAgICAgICAgbmFtZSA9IG5hbWUucmVwbGFjZShjaCwgJycpCiAgICByZXR1cm4gbmFtZVs6bWF4X2xlbl0KCgojIOKUgOKUgCBQYXJzZSB0aGUgY29uc29saWRhdGVkIENTViDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZGVmIHBhcnNlX2Nzdihjc3ZfcGF0aCk6CiAgICAiIiJSZXR1cm4gbWV0YWRhdGEgZGljdCwgaGVhZGVyIGxpc3QsIGFuZCBkYXRhIHJvd3MuIiIiCiAgICBtZXRhID0ge30KICAgIGhlYWRlciA9IE5vbmUKICAgIHJvd3MgPSBbXQoKICAgIHdpdGggb3Blbihjc3ZfcGF0aCwgInIiKSBhcyBmOgogICAgICAgIGZvciBsaW5lIGluIGY6CiAgICAgICAgICAgIGxpbmUgPSBsaW5lLnN0cmlwKCkKICAgICAgICAgICAgaWYgbm90IGxpbmU6CiAgICAgICAgICAgICAgICBjb250aW51ZQogICAgICAgICAgICBpZiBsaW5lLnN0YXJ0c3dpdGgoIiMiKToKICAgICAgICAgICAgICAgIGNvbnRlbnQgPSBsaW5lLmxzdHJpcCgiIyAiKS5zdHJpcCgpCiAgICAgICAgICAgICAgICBpZiAiOiIgaW4gY29udGVudDoKICAgICAgICAgICAgICAgICAgICBrZXksIHZhbCA9IGNvbnRlbnQuc3BsaXQoIjoiLCAxKQogICAgICAgICAgICAgICAgICAgIG1ldGFba2V5LnN0cmlwKCldID0gdmFsLnN0cmlwKCkKICAgICAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgICAgIGlmIGhlYWRlciBpcyBOb25lOgogICAgICAgICAgICAgICAgaGVhZGVyID0gW2guc3RyaXAoKSBmb3IgaCBpbiBsaW5lLnNwbGl0KCIsIildCiAgICAgICAgICAgICAgICBjb250aW51ZQogICAgICAgICAgICB2YWx1ZXMgPSBbdi5zdHJpcCgpIGZvciB2IGluIGxpbmUuc3BsaXQoIiwiKV0KICAgICAgICAgICAgcm93ID0ge30KICAgICAgICAgICAgZm9yIGksIGggaW4gZW51bWVyYXRlKGhlYWRlcik6CiAgICAgICAgICAgICAgICByb3dbaF0gPSB2YWx1ZXNbaV0gaWYgaSA8IGxlbih2YWx1ZXMpIGVsc2UgIiIKICAgICAgICAgICAgcm93cy5hcHBlbmQocm93KQoKICAgIHJldHVybiBtZXRhLCBoZWFkZXIsIHJvd3MKCgpkZWYgZGV0ZWN0X3ZhcmlhYmxlX2xhYmVscyhoZWFkZXIpOgogICAgIiIiCiAgICBFeHRyYWN0IGJhc2UgdmFyaWFibGUgbGFiZWxzIGZyb20gaGVhZGVyLgogICAgZS5nLiBoZWFkZXIgaGFzIExFMjJfTWF4LCBMRTIyX01pbiwgUzIyX01heCwgUzIyX01pbgogICAgcmV0dXJucyBbJ0xFMjInLCAnUzIyJ10gaW4gb3JkZXIKICAgICIiIgogICAgdmFyX3N0YXJ0ID0gaGVhZGVyLmluZGV4KCJZTGFiZWwiKSArIDEKICAgIHJhd19jb2xzID0gaGVhZGVyW3Zhcl9zdGFydDpdCiAgICBsYWJlbHMgPSBbXQogICAgc2VlbiA9IHNldCgpCiAgICBmb3IgY29sIGluIHJhd19jb2xzOgogICAgICAgICMgU3RyaXAgX01heCBvciBfTWluIHN1ZmZpeAogICAgICAgIGlmIGNvbC5lbmRzd2l0aCgiX01heCIpOgogICAgICAgICAgICBiYXNlID0gY29sWzotNF0KICAgICAgICBlbGlmIGNvbC5lbmRzd2l0aCgiX01pbiIpOgogICAgICAgICAgICBiYXNlID0gY29sWzotNF0KICAgICAgICBlbHNlOgogICAgICAgICAgICBiYXNlID0gY29sCiAgICAgICAgaWYgYmFzZSBub3QgaW4gc2VlbjoKICAgICAgICAgICAgc2Vlbi5hZGQoYmFzZSkKICAgICAgICAgICAgbGFiZWxzLmFwcGVuZChiYXNlKQogICAgcmV0dXJuIGxhYmVscwoKCmRlZiBidWlsZF9kYXRhKHJvd3MsIHZhcl9sYWJlbHMpOgogICAgIiIiCiAgICBCdWlsZCBzdHJ1Y3R1cmVkIGRhdGEgZGljdC4KCiAgICBSZXR1cm5zOgogICAgICAgIHhfbGFiZWxzOiAgIHNvcnRlZCB1bmlxdWUgWCBwb3NpdGlvbiBsYWJlbHMKICAgICAgICB5X2xhYmVsczogICBzb3J0ZWQgdW5pcXVlIFkgcG9zaXRpb24gbGFiZWxzCiAgICAgICAgc3RlcF9vcmRlcjogbGlzdCBvZiBzdGVwIG5hbWVzIGluIG9yZGVyIG9mIGFwcGVhcmFuY2UKICAgICAgICBkYXRhOiAgICAgICB7KHN0ZXAsIHhpLCB5aSk6IHt2YXJfbGFiZWw6IChtYXhfdmFsLCBtaW5fdmFsKX19CiAgICAiIiIKICAgIHhfc2V0ID0gc2V0KCkKICAgIHlfc2V0ID0gc2V0KCkKICAgIHN0ZXBfb3JkZXIgPSBbXQogICAgc3RlcF9zZWVuID0gc2V0KCkKICAgIGRhdGEgPSB7fQoKICAgIGZvciByb3cgaW4gcm93czoKICAgICAgICBzdGVwID0gcm93LmdldCgiU3RlcCIsICIiKQogICAgICAgIHhpID0gaW50KHJvdy5nZXQoIlhpIiwgMCkpCiAgICAgICAgeWkgPSBpbnQocm93LmdldCgiWWkiLCAwKSkKICAgICAgICB4X2xhYmVsID0gcm93LmdldCgiWExhYmVsIiwgIiIpCiAgICAgICAgeV9sYWJlbCA9IHJvdy5nZXQoIllMYWJlbCIsICIiKQoKICAgICAgICB4X3NldC5hZGQoKHhpLCB4X2xhYmVsKSkKICAgICAgICB5X3NldC5hZGQoKHlpLCB5X2xhYmVsKSkKCiAgICAgICAgaWYgc3RlcCBub3QgaW4gc3RlcF9zZWVuOgogICAgICAgICAgICBzdGVwX3NlZW4uYWRkKHN0ZXApCiAgICAgICAgICAgIHN0ZXBfb3JkZXIuYXBwZW5kKHN0ZXApCgogICAgICAgIGtleSA9IChzdGVwLCB4aSwgeWkpCiAgICAgICAgaWYga2V5IG5vdCBpbiBkYXRhOgogICAgICAgICAgICBkYXRhW2tleV0gPSB7fQoKICAgICAgICBmb3IgbGFiZWwgaW4gdmFyX2xhYmVsczoKICAgICAgICAgICAgbWF4X2NvbCA9ICIlc19NYXgiICUgbGFiZWwKICAgICAgICAgICAgbWluX2NvbCA9ICIlc19NaW4iICUgbGFiZWwKICAgICAgICAgICAgbWF4X3ZhbCA9IHJvdy5nZXQobWF4X2NvbCwgIk4vQSIpCiAgICAgICAgICAgIG1pbl92YWwgPSByb3cuZ2V0KG1pbl9jb2wsICJOL0EiKQogICAgICAgICAgICBkYXRhW2tleV1bbGFiZWxdID0gKG1heF92YWwsIG1pbl92YWwpCgogICAgIyBTb3J0IGJ5IGluZGV4CiAgICB4X2xhYmVscyA9IFtsYmwgZm9yIF8sIGxibCBpbiBzb3J0ZWQoeF9zZXQpXQogICAgeV9sYWJlbHMgPSBbbGJsIGZvciBfLCBsYmwgaW4gc29ydGVkKHlfc2V0KV0KCiAgICByZXR1cm4geF9sYWJlbHMsIHlfbGFiZWxzLCBzdGVwX29yZGVyLCBkYXRhCgoKIyDilIDilIAgV3JpdGUgU3VtbWFyeSBzaGVldCDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZGVmIHdyaXRlX3N1bW1hcnkod2IsIHNldF9uYW1lLCB2YXJfbGFiZWxzLCB4X2xhYmVscywgeV9sYWJlbHMsIHN0ZXBfb3JkZXIsCiAgICAgICAgICAgICAgICAgIGxvYWRpbmdfc3RlcHMsIHJlbGF4YXRpb25fc3RlcHMsIGRhdGEpOgogICAgd3MgPSB3Yi5jcmVhdGVfc2hlZXQodGl0bGU9IlN1bW1hcnkiKQoKICAgIHdzLmNlbGwocm93PTEsIGNvbHVtbj0xLCB2YWx1ZT0iRWxlbWVudCBTZXQ6ICVzIiAlIHNldF9uYW1lKQogICAgd3MuY2VsbChyb3c9MSwgY29sdW1uPTEpLmZvbnQgPSBGb250KGJvbGQ9VHJ1ZSwgc2l6ZT0xNCwgY29sb3I9IkZGRkZGRiIpCiAgICB3cy5jZWxsKHJvdz0xLCBjb2x1bW49MSkuZmlsbCA9IEhFQURFUl9GSUxMX0RBUksKCiAgICB3cy5jZWxsKHJvdz0yLCBjb2x1bW49MSwKICAgICAgICAgICAgdmFsdWU9IlN0ZXBzOiAlZCAgKCVkIExvYWRpbmcgKyAlZCBSZWxheGF0aW9uKSIgJSAoCiAgICAgICAgICAgICAgICBsZW4oc3RlcF9vcmRlciksIGxlbihsb2FkaW5nX3N0ZXBzKSwgbGVuKHJlbGF4YXRpb25fc3RlcHMpKSkKICAgIHdzLmNlbGwocm93PTIsIGNvbHVtbj0xKS5mb250ID0gRm9udChzaXplPTEwLCBjb2xvcj0iQTBBRUMwIikKCiAgICB3cy5jZWxsKHJvdz0zLCBjb2x1bW49MSwKICAgICAgICAgICAgdmFsdWU9IlJlZ2lvbnM6ICVkICAoJWQgWCB4ICVkIFkpIiAlICgKICAgICAgICAgICAgICAgIGxlbih4X2xhYmVscykgKiBtYXgobGVuKHlfbGFiZWxzKSwgMSksCiAgICAgICAgICAgICAgICBsZW4oeF9sYWJlbHMpLCBtYXgobGVuKHlfbGFiZWxzKSwgMSkpKQogICAgd3MuY2VsbChyb3c9MywgY29sdW1uPTEpLmZvbnQgPSBGb250KHNpemU9MTAsIGNvbG9yPSJBMEFFQzAiKQoKICAgICMgUGVhay12YWx1ZXMgdGFibGUKICAgIHN1bV9yb3cgPSA1CiAgICB3cy5jZWxsKHJvdz1zdW1fcm93LCBjb2x1bW49MSwgdmFsdWU9IlN0ZXAiKQogICAgd3MuY2VsbChyb3c9c3VtX3JvdywgY29sdW1uPTIsIHZhbHVlPSJQaGFzZSIpCiAgICBjb2xfaWR4ID0gMwogICAgZm9yIGxhYmVsIGluIHZhcl9sYWJlbHM6CiAgICAgICAgd3MuY2VsbChyb3c9c3VtX3JvdywgY29sdW1uPWNvbF9pZHgsIHZhbHVlPSIlcyBNYXgiICUgbGFiZWwpCiAgICAgICAgd3MuY2VsbChyb3c9c3VtX3JvdywgY29sdW1uPWNvbF9pZHggKyAxLCB2YWx1ZT0iJXMgTWluIiAlIGxhYmVsKQogICAgICAgIGNvbF9pZHggKz0gMgogICAgX3N0eWxlX2hlYWRlcl9yb3cod3MsIHN1bV9yb3csIGNvbF9pZHggLSAxLCBTVU1NQVJZX0hFQURFUl9GSUxMLCBTVU1NQVJZX0hFQURFUl9GT05UKQoKICAgIGZvciBzdGVwX2lkeCwgc3RlcCBpbiBlbnVtZXJhdGUoc3RlcF9vcmRlcik6CiAgICAgICAgciA9IHN1bV9yb3cgKyBzdGVwX2lkeCArIDEKICAgICAgICBwaGFzZSA9ICJMb2FkaW5nIiBpZiBzdGVwIGluIGxvYWRpbmdfc3RlcHMgZWxzZSAiUmVsYXhhdGlvbiIKICAgICAgICB3cy5jZWxsKHJvdz1yLCBjb2x1bW49MSwgdmFsdWU9c3RlcCkuYm9yZGVyID0gVEhJTl9CT1JERVIKICAgICAgICBwaGFzZV9jZWxsID0gd3MuY2VsbChyb3c9ciwgY29sdW1uPTIsIHZhbHVlPXBoYXNlKQogICAgICAgIHBoYXNlX2NlbGwuYm9yZGVyID0gVEhJTl9CT1JERVIKICAgICAgICBpZiBwaGFzZSA9PSAiTG9hZGluZyI6CiAgICAgICAgICAgIHBoYXNlX2NlbGwuZm9udCA9IExPQURJTkdfRk9OVAogICAgICAgICAgICBwaGFzZV9jZWxsLmZpbGwgPSBMT0FESU5HX0ZJTEwKICAgICAgICBlbHNlOgogICAgICAgICAgICBwaGFzZV9jZWxsLmZvbnQgPSBSRUxBWF9GT05UCiAgICAgICAgICAgIHBoYXNlX2NlbGwuZmlsbCA9IFJFTEFYX0ZJTEwKCiAgICAgICAgY29sX2lkeCA9IDMKICAgICAgICBmb3IgbGFiZWwgaW4gdmFyX2xhYmVsczoKICAgICAgICAgICAgcGVha19tYXggPSBOb25lCiAgICAgICAgICAgIHBlYWtfbWluID0gTm9uZQogICAgICAgICAgICBmb3IgeGkgaW4gcmFuZ2UobGVuKHhfbGFiZWxzKSk6CiAgICAgICAgICAgICAgICBmb3IgeWkgaW4gcmFuZ2UobGVuKHlfbGFiZWxzKSk6CiAgICAgICAgICAgICAgICAgICAga2V5ID0gKHN0ZXAsIHhpLCB5aSkKICAgICAgICAgICAgICAgICAgICBpZiBrZXkgaW4gZGF0YSBhbmQgbGFiZWwgaW4gZGF0YVtrZXldOgogICAgICAgICAgICAgICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICBtdiA9IGZsb2F0KGRhdGFba2V5XVtsYWJlbF1bMF0pCiAgICAgICAgICAgICAgICAgICAgICAgICAgICBpZiBwZWFrX21heCBpcyBOb25lIG9yIG12ID4gcGVha19tYXg6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgcGVha19tYXggPSBtdgogICAgICAgICAgICAgICAgICAgICAgICBleGNlcHQgKFZhbHVlRXJyb3IsIFR5cGVFcnJvcik6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICBwYXNzCiAgICAgICAgICAgICAgICAgICAgICAgIHRyeToKICAgICAgICAgICAgICAgICAgICAgICAgICAgIG52ID0gZmxvYXQoZGF0YVtrZXldW2xhYmVsXVsxXSkKICAgICAgICAgICAgICAgICAgICAgICAgICAgIGlmIHBlYWtfbWluIGlzIE5vbmUgb3IgbnYgPCBwZWFrX21pbjoKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBwZWFrX21pbiA9IG52CiAgICAgICAgICAgICAgICAgICAgICAgIGV4Y2VwdCAoVmFsdWVFcnJvciwgVHlwZUVycm9yKToKICAgICAgICAgICAgICAgICAgICAgICAgICAgIHBhc3MKCiAgICAgICAgICAgIGlmIHBlYWtfbWF4IGlzIG5vdCBOb25lOgogICAgICAgICAgICAgICAgX3dyaXRlX3ZhbHVlKHdzLCByLCBjb2xfaWR4LCAiJS44ZSIgJSBwZWFrX21heCkKICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgIHdzLmNlbGwocm93PXIsIGNvbHVtbj1jb2xfaWR4LCB2YWx1ZT0iTi9BIikuYm9yZGVyID0gVEhJTl9CT1JERVIKICAgICAgICAgICAgaWYgcGVha19taW4gaXMgbm90IE5vbmU6CiAgICAgICAgICAgICAgICBfd3JpdGVfdmFsdWUod3MsIHIsIGNvbF9pZHggKyAxLCAiJS44ZSIgJSBwZWFrX21pbikKICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgIHdzLmNlbGwocm93PXIsIGNvbHVtbj1jb2xfaWR4ICsgMSwgdmFsdWU9Ik4vQSIpLmJvcmRlciA9IFRISU5fQk9SREVSCiAgICAgICAgICAgIGNvbF9pZHggKz0gMgoKICAgIF9hdXRvX2ZpdCh3cykKCgojIOKUgOKUgCBXcml0ZSBwZXItdmFyaWFibGUgZGF0YSBzaGVldHMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmRlZiB3cml0ZV92YXJpYWJsZV9zaGVldHMod2IsIHZhcl9sYWJlbHMsIHhfbGFiZWxzLCB5X2xhYmVscywgc3RlcF9vcmRlciwKICAgICAgICAgICAgICAgICAgICAgICAgICBsb2FkaW5nX3N0ZXBzLCByZWxheGF0aW9uX3N0ZXBzLCBkYXRhLCBpc18yZCwgc2V0X25hbWUpOgogICAgcGhhc2VzID0gWwogICAgICAgICgiTG9hZGluZyIsIGxvYWRpbmdfc3RlcHMsIExPQURJTkdfRklMTCwgTE9BRElOR19GT05UKSwKICAgICAgICAoIlJlbGF4IiwgcmVsYXhhdGlvbl9zdGVwcywgUkVMQVhfRklMTCwgUkVMQVhfRk9OVCksCiAgICBdCgogICAgcG9zX2xhYmVscyA9IHhfbGFiZWxzIGlmIGxlbih4X2xhYmVscykgPiAxIGVsc2UgeV9sYWJlbHMKICAgIHBvc19heGlzID0gIlgiIGlmIGxlbih4X2xhYmVscykgPiAxIGVsc2UgIlkiCgogICAgZm9yIGxhYmVsIGluIHZhcl9sYWJlbHM6CiAgICAgICAgZm9yIHN0YXRfdHlwZSwgc3RhdF9pZHggaW4gWygiTWF4IiwgMCksICgiTWluIiwgMSldOgogICAgICAgICAgICAjIOKUgOKUgCBQaGFzZS1zcGVjaWZpYyBzaGVldHMgKExvYWRpbmcgLyBSZWxheGF0aW9uKSDilIDilIAKICAgICAgICAgICAgZm9yIHBoYXNlX25hbWUsIHBoYXNlX3N0ZXBzLCBwaGFzZV9maWxsLCBwaGFzZV9mb250IGluIHBoYXNlczoKICAgICAgICAgICAgICAgIGlmIG5vdCBwaGFzZV9zdGVwczoKICAgICAgICAgICAgICAgICAgICBjb250aW51ZQoKICAgICAgICAgICAgICAgIHNoZWV0X25hbWUgPSBfc2FmZV9zaGVldF9uYW1lKCIlc18lc18lcyIgJSAobGFiZWwsIHN0YXRfdHlwZSwgcGhhc2VfbmFtZSkpCiAgICAgICAgICAgICAgICB3cyA9IHdiLmNyZWF0ZV9zaGVldCh0aXRsZT1zaGVldF9uYW1lKQoKICAgICAgICAgICAgICAgIGlmIGlzXzJkOgogICAgICAgICAgICAgICAgICAgIF93cml0ZV8yZF9zaGVldCh3cywgbGFiZWwsIHN0YXRfdHlwZSwgc3RhdF9pZHgsIHBoYXNlX25hbWUsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgcGhhc2Vfc3RlcHMsIHBoYXNlX2ZpbGwsIHBoYXNlX2ZvbnQsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgeF9sYWJlbHMsIHlfbGFiZWxzLCBkYXRhKQogICAgICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgICAgICBfd3JpdGVfMWRfc2hlZXQod3MsIGxhYmVsLCBzdGF0X3R5cGUsIHN0YXRfaWR4LCBwaGFzZV9uYW1lLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIHBoYXNlX3N0ZXBzLCBwaGFzZV9maWxsLCBwaGFzZV9mb250LAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIHBvc19sYWJlbHMsIHBvc19heGlzLCBkYXRhLCBzZXRfbmFtZSkKICAgICAgICAgICAgICAgIF9hdXRvX2ZpdCh3cykKCiAgICAgICAgICAgICMg4pSA4pSAIEFsbFN0ZXBzIHNoZWV0IOKUgOKUgAogICAgICAgICAgICBpZiBsZW4oc3RlcF9vcmRlcikgPiAxOgogICAgICAgICAgICAgICAgc2hlZXRfbmFtZSA9IF9zYWZlX3NoZWV0X25hbWUoIiVzXyVzX0FsbFN0ZXBzIiAlIChsYWJlbCwgc3RhdF90eXBlKSkKICAgICAgICAgICAgICAgIHdzID0gd2IuY3JlYXRlX3NoZWV0KHRpdGxlPXNoZWV0X25hbWUpCgogICAgICAgICAgICAgICAgaWYgaXNfMmQ6CiAgICAgICAgICAgICAgICAgICAgY3VycmVudF9yb3cgPSAxCiAgICAgICAgICAgICAgICAgICAgZm9yIHN0ZXAgaW4gc3RlcF9vcmRlcjoKICAgICAgICAgICAgICAgICAgICAgICAgcGhhc2UgPSAiTG9hZGluZyIgaWYgc3RlcCBpbiBsb2FkaW5nX3N0ZXBzIGVsc2UgIlJlbGF4YXRpb24iCiAgICAgICAgICAgICAgICAgICAgICAgIHBmID0gTE9BRElOR19GSUxMIGlmIHBoYXNlID09ICJMb2FkaW5nIiBlbHNlIFJFTEFYX0ZJTEwKICAgICAgICAgICAgICAgICAgICAgICAgcGZvbnQgPSBMT0FESU5HX0ZPTlQgaWYgcGhhc2UgPT0gIkxvYWRpbmciIGVsc2UgUkVMQVhfRk9OVAoKICAgICAgICAgICAgICAgICAgICAgICAgd3MuY2VsbChyb3c9Y3VycmVudF9yb3csIGNvbHVtbj0xLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIHZhbHVlPSIlcyAgfCAgJXMgJXMgIHwgICVzIiAlIChzdGVwLCBsYWJlbCwgc3RhdF90eXBlLCBwaGFzZSkpCiAgICAgICAgICAgICAgICAgICAgICAgIHdzLmNlbGwocm93PWN1cnJlbnRfcm93LCBjb2x1bW49MSkuZm9udCA9IHBmb250CiAgICAgICAgICAgICAgICAgICAgICAgIHdzLmNlbGwocm93PWN1cnJlbnRfcm93LCBjb2x1bW49MSkuZmlsbCA9IHBmCiAgICAgICAgICAgICAgICAgICAgICAgIHdzLm1lcmdlX2NlbGxzKHN0YXJ0X3Jvdz1jdXJyZW50X3Jvdywgc3RhcnRfY29sdW1uPTEsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIGVuZF9yb3c9Y3VycmVudF9yb3csIGVuZF9jb2x1bW49bGVuKHhfbGFiZWxzKSArIDEpCiAgICAgICAgICAgICAgICAgICAgICAgIGN1cnJlbnRfcm93ICs9IDEKCiAgICAgICAgICAgICAgICAgICAgICAgIHdzLmNlbGwocm93PWN1cnJlbnRfcm93LCBjb2x1bW49MSwgdmFsdWU9IlkgXFwgWCIpCiAgICAgICAgICAgICAgICAgICAgICAgIGZvciBjaSwgeGwgaW4gZW51bWVyYXRlKHhfbGFiZWxzKToKICAgICAgICAgICAgICAgICAgICAgICAgICAgIHdzLmNlbGwocm93PWN1cnJlbnRfcm93LCBjb2x1bW49Y2kgKyAyLCB2YWx1ZT14bCkKICAgICAgICAgICAgICAgICAgICAgICAgX3N0eWxlX2hlYWRlcl9yb3cod3MsIGN1cnJlbnRfcm93LCBsZW4oeF9sYWJlbHMpICsgMSkKICAgICAgICAgICAgICAgICAgICAgICAgY3VycmVudF9yb3cgKz0gMQoKICAgICAgICAgICAgICAgICAgICAgICAgZm9yIHlpX2lkeCwgeWwgaW4gZW51bWVyYXRlKHlfbGFiZWxzKToKICAgICAgICAgICAgICAgICAgICAgICAgICAgIF93cml0ZV9wb3NfbGFiZWwod3MsIGN1cnJlbnRfcm93LCAxLCB5bCkKICAgICAgICAgICAgICAgICAgICAgICAgICAgIGZvciB4aV9pZHggaW4gcmFuZ2UobGVuKHhfbGFiZWxzKSk6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAga2V5ID0gKHN0ZXAsIHhpX2lkeCwgeWlfaWR4KQogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIGlmIGtleSBpbiBkYXRhIGFuZCBsYWJlbCBpbiBkYXRhW2tleV06CiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIF93cml0ZV92YWx1ZSh3cywgY3VycmVudF9yb3csIHhpX2lkeCArIDIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBkYXRhW2tleV1bbGFiZWxdW3N0YXRfaWR4XSkKICAgICAgICAgICAgICAgICAgICAgICAgICAgIGN1cnJlbnRfcm93ICs9IDEKICAgICAgICAgICAgICAgICAgICAgICAgY3VycmVudF9yb3cgKz0gMSAgIyBibGFuayBzZXBhcmF0b3Igcm93CgogICAgICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgICAgICAjIDFEOiByb3dzPXBvc2l0aW9ucywgY29sdW1ucz1BTEwgc3RlcCBuYW1lcyAoY29sb3ItY29kZWQpCiAgICAgICAgICAgICAgICAgICAgd3MuY2VsbChyb3c9MSwgY29sdW1uPTEsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICB2YWx1ZT0iJXMgIHwgICVzICVzICB8ICBBbGwgU3RlcHMiICUgKHNldF9uYW1lLCBsYWJlbCwgc3RhdF90eXBlKSkKICAgICAgICAgICAgICAgICAgICB3cy5jZWxsKHJvdz0xLCBjb2x1bW49MSkuZm9udCA9IEZvbnQoYm9sZD1UcnVlLCBzaXplPTEyLCBjb2xvcj0iRkZGRkZGIikKICAgICAgICAgICAgICAgICAgICB3cy5tZXJnZV9jZWxscyhzdGFydF9yb3c9MSwgc3RhcnRfY29sdW1uPTEsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgZW5kX3Jvdz0xLCBlbmRfY29sdW1uPWxlbihzdGVwX29yZGVyKSArIDEpCiAgICAgICAgICAgICAgICAgICAgd3MuY2VsbChyb3c9MSwgY29sdW1uPTEpLmZpbGwgPSBIRUFERVJfRklMTF9EQVJLCgogICAgICAgICAgICAgICAgICAgIHdzLmNlbGwocm93PTIsIGNvbHVtbj0xLCB2YWx1ZT0iUG9zaXRpb25fJXMgKG1tKSIgJSBwb3NfYXhpcykKICAgICAgICAgICAgICAgICAgICBmb3Igc2ksIHN0ZXAgaW4gZW51bWVyYXRlKHN0ZXBfb3JkZXIpOgogICAgICAgICAgICAgICAgICAgICAgICBjZWxsID0gd3MuY2VsbChyb3c9MiwgY29sdW1uPXNpICsgMiwgdmFsdWU9c3RlcCkKICAgICAgICAgICAgICAgICAgICAgICAgY2VsbC5ib3JkZXIgPSBUSElOX0JPUkRFUgogICAgICAgICAgICAgICAgICAgICAgICBjZWxsLmFsaWdubWVudCA9IEFsaWdubWVudChob3Jpem9udGFsPSJjZW50ZXIiKQogICAgICAgICAgICAgICAgICAgICAgICBpZiBzdGVwIGluIGxvYWRpbmdfc3RlcHM6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICBjZWxsLmZvbnQgPSBMT0FESU5HX0ZPTlQKICAgICAgICAgICAgICAgICAgICAgICAgICAgIGNlbGwuZmlsbCA9IExPQURJTkdfRklMTAogICAgICAgICAgICAgICAgICAgICAgICBlbHNlOgogICAgICAgICAgICAgICAgICAgICAgICAgICAgY2VsbC5mb250ID0gUkVMQVhfRk9OVAogICAgICAgICAgICAgICAgICAgICAgICAgICAgY2VsbC5maWxsID0gUkVMQVhfRklMTAoKICAgICAgICAgICAgICAgICAgICBwb3NfY2VsbCA9IHdzLmNlbGwocm93PTIsIGNvbHVtbj0xKQogICAgICAgICAgICAgICAgICAgIHBvc19jZWxsLmZvbnQgPSBIRUFERVJfRk9OVF9XSElURQogICAgICAgICAgICAgICAgICAgIHBvc19jZWxsLmZpbGwgPSBIRUFERVJfRklMTF9EQVJLCiAgICAgICAgICAgICAgICAgICAgcG9zX2NlbGwuYm9yZGVyID0gVEhJTl9CT1JERVIKCiAgICAgICAgICAgICAgICAgICAgZm9yIGNpLCBwbCBpbiBlbnVtZXJhdGUocG9zX2xhYmVscyk6CiAgICAgICAgICAgICAgICAgICAgICAgIHJvd19udW0gPSBjaSArIDMKICAgICAgICAgICAgICAgICAgICAgICAgX3dyaXRlX3Bvc19sYWJlbCh3cywgcm93X251bSwgMSwgcGwpCiAgICAgICAgICAgICAgICAgICAgICAgIGZvciBzaSwgc3RlcCBpbiBlbnVtZXJhdGUoc3RlcF9vcmRlcik6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICBpZiBwb3NfYXhpcyA9PSAiWCI6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAga2V5ID0gKHN0ZXAsIGNpLCAwKQogICAgICAgICAgICAgICAgICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBrZXkgPSAoc3RlcCwgMCwgY2kpCiAgICAgICAgICAgICAgICAgICAgICAgICAgICBpZiBrZXkgaW4gZGF0YSBhbmQgbGFiZWwgaW4gZGF0YVtrZXldOgogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIF93cml0ZV92YWx1ZSh3cywgcm93X251bSwgc2kgKyAyLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBkYXRhW2tleV1bbGFiZWxdW3N0YXRfaWR4XSkKCiAgICAgICAgICAgICAgICBfYXV0b19maXQod3MpCgoKZGVmIF93cml0ZV8xZF9zaGVldCh3cywgbGFiZWwsIHN0YXRfdHlwZSwgc3RhdF9pZHgsIHBoYXNlX25hbWUsCiAgICAgICAgICAgICAgICAgICAgcGhhc2Vfc3RlcHMsIHBoYXNlX2ZpbGwsIHBoYXNlX2ZvbnQsCiAgICAgICAgICAgICAgICAgICAgcG9zX2xhYmVscywgcG9zX2F4aXMsIGRhdGEsIHNldF9uYW1lKToKICAgICIiIjFEIHN3ZWVwOiByb3dzID0gcG9zaXRpb25zLCBjb2x1bW5zID0gc3RlcCBuYW1lcy4iIiIKICAgICMgVGl0bGUgcm93CiAgICB3cy5jZWxsKHJvdz0xLCBjb2x1bW49MSwKICAgICAgICAgICAgdmFsdWU9IiVzICB8ICAlcyAlcyAgfCAgJXMiICUgKHNldF9uYW1lLCBsYWJlbCwgc3RhdF90eXBlLCBwaGFzZV9uYW1lKSkKICAgIHdzLmNlbGwocm93PTEsIGNvbHVtbj0xKS5mb250ID0gRm9udChib2xkPVRydWUsIHNpemU9MTIsIGNvbG9yPSJGRkZGRkYiKQogICAgd3MubWVyZ2VfY2VsbHMoc3RhcnRfcm93PTEsIHN0YXJ0X2NvbHVtbj0xLAogICAgICAgICAgICAgICAgICAgZW5kX3Jvdz0xLCBlbmRfY29sdW1uPWxlbihwaGFzZV9zdGVwcykgKyAxKQogICAgd3MuY2VsbChyb3c9MSwgY29sdW1uPTEpLmZpbGwgPSBwaGFzZV9maWxsCgogICAgIyBIZWFkZXI6IFBvc2l0aW9uIHwgU3RlcDEgfCBTdGVwMiB8IC4uLgogICAgd3MuY2VsbChyb3c9MiwgY29sdW1uPTEsIHZhbHVlPSJQb3NpdGlvbl8lcyAobW0pIiAlIHBvc19heGlzKQogICAgZm9yIHNpLCBzdGVwIGluIGVudW1lcmF0ZShwaGFzZV9zdGVwcyk6CiAgICAgICAgd3MuY2VsbChyb3c9MiwgY29sdW1uPXNpICsgMiwgdmFsdWU9c3RlcCkKICAgIF9zdHlsZV9oZWFkZXJfcm93KHdzLCAyLCBsZW4ocGhhc2Vfc3RlcHMpICsgMSwgcGhhc2VfZmlsbCwgcGhhc2VfZm9udCkKCiAgICAjIERhdGEgcm93cwogICAgZm9yIGNpLCBwbCBpbiBlbnVtZXJhdGUocG9zX2xhYmVscyk6CiAgICAgICAgcm93X251bSA9IGNpICsgMwogICAgICAgIF93cml0ZV9wb3NfbGFiZWwod3MsIHJvd19udW0sIDEsIHBsKQogICAgICAgIGZvciBzaSwgc3RlcCBpbiBlbnVtZXJhdGUocGhhc2Vfc3RlcHMpOgogICAgICAgICAgICBpZiBwb3NfYXhpcyA9PSAiWCI6CiAgICAgICAgICAgICAgICBrZXkgPSAoc3RlcCwgY2ksIDApCiAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICBrZXkgPSAoc3RlcCwgMCwgY2kpCiAgICAgICAgICAgIGlmIGtleSBpbiBkYXRhIGFuZCBsYWJlbCBpbiBkYXRhW2tleV06CiAgICAgICAgICAgICAgICBfd3JpdGVfdmFsdWUod3MsIHJvd19udW0sIHNpICsgMiwgZGF0YVtrZXldW2xhYmVsXVtzdGF0X2lkeF0pCgoKZGVmIF93cml0ZV8yZF9zaGVldCh3cywgbGFiZWwsIHN0YXRfdHlwZSwgc3RhdF9pZHgsIHBoYXNlX25hbWUsCiAgICAgICAgICAgICAgICAgICAgcGhhc2Vfc3RlcHMsIHBoYXNlX2ZpbGwsIHBoYXNlX2ZvbnQsCiAgICAgICAgICAgICAgICAgICAgeF9sYWJlbHMsIHlfbGFiZWxzLCBkYXRhKToKICAgICIiIjJEIHN3ZWVwOiBzdGFja2VkIFkgeCBYIGJsb2Nrcywgb25lIHBlciBzdGVwLiIiIgogICAgY3VycmVudF9yb3cgPSAxCiAgICBmb3Igc3RlcCBpbiBwaGFzZV9zdGVwczoKICAgICAgICAjIFN0ZXAgaGVhZGVyCiAgICAgICAgd3MuY2VsbChyb3c9Y3VycmVudF9yb3csIGNvbHVtbj0xLAogICAgICAgICAgICAgICAgdmFsdWU9IiVzICB8ICAlcyAlcyAgfCAgJXMiICUgKHN0ZXAsIGxhYmVsLCBzdGF0X3R5cGUsIHBoYXNlX25hbWUpKQogICAgICAgIHdzLmNlbGwocm93PWN1cnJlbnRfcm93LCBjb2x1bW49MSkuZm9udCA9IHBoYXNlX2ZvbnQKICAgICAgICB3cy5jZWxsKHJvdz1jdXJyZW50X3JvdywgY29sdW1uPTEpLmZpbGwgPSBwaGFzZV9maWxsCiAgICAgICAgd3MubWVyZ2VfY2VsbHMoc3RhcnRfcm93PWN1cnJlbnRfcm93LCBzdGFydF9jb2x1bW49MSwKICAgICAgICAgICAgICAgICAgICAgICBlbmRfcm93PWN1cnJlbnRfcm93LCBlbmRfY29sdW1uPWxlbih4X2xhYmVscykgKyAxKQogICAgICAgIGN1cnJlbnRfcm93ICs9IDEKCiAgICAgICAgIyBDb2x1bW4gaGVhZGVycyAoWCBwb3NpdGlvbnMpCiAgICAgICAgd3MuY2VsbChyb3c9Y3VycmVudF9yb3csIGNvbHVtbj0xLCB2YWx1ZT0iWSBcXCBYIikKICAgICAgICBmb3IgY2ksIHhsIGluIGVudW1lcmF0ZSh4X2xhYmVscyk6CiAgICAgICAgICAgIHdzLmNlbGwocm93PWN1cnJlbnRfcm93LCBjb2x1bW49Y2kgKyAyLCB2YWx1ZT14bCkKICAgICAgICBfc3R5bGVfaGVhZGVyX3Jvdyh3cywgY3VycmVudF9yb3csIGxlbih4X2xhYmVscykgKyAxKQogICAgICAgIGN1cnJlbnRfcm93ICs9IDEKCiAgICAgICAgIyBEYXRhIHJvd3MgKG9uZSBwZXIgWSBwb3NpdGlvbikKICAgICAgICBmb3IgeWlfaWR4LCB5bCBpbiBlbnVtZXJhdGUoeV9sYWJlbHMpOgogICAgICAgICAgICBfd3JpdGVfcG9zX2xhYmVsKHdzLCBjdXJyZW50X3JvdywgMSwgeWwpCiAgICAgICAgICAgIGZvciB4aV9pZHggaW4gcmFuZ2UobGVuKHhfbGFiZWxzKSk6CiAgICAgICAgICAgICAgICBrZXkgPSAoc3RlcCwgeGlfaWR4LCB5aV9pZHgpCiAgICAgICAgICAgICAgICBpZiBrZXkgaW4gZGF0YSBhbmQgbGFiZWwgaW4gZGF0YVtrZXldOgogICAgICAgICAgICAgICAgICAgIF93cml0ZV92YWx1ZSh3cywgY3VycmVudF9yb3csIHhpX2lkeCArIDIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIGRhdGFba2V5XVtsYWJlbF1bc3RhdF9pZHhdKQogICAgICAgICAgICBjdXJyZW50X3JvdyArPSAxCgogICAgICAgIGN1cnJlbnRfcm93ICs9IDEgICMgYmxhbmsgcm93IGJldHdlZW4gc3RlcCBibG9ja3MKCgojIOKUgOKUgCBNYWluIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApkZWYgY29udmVydF9jc3ZfdG9fZXhjZWwoY3N2X3BhdGgsIHhsc3hfcGF0aD1Ob25lKToKICAgICIiIkNvbnZlcnQgYSBjb25zb2xpZGF0ZWQgQ1NWIHRvIGEgZm9ybWF0dGVkLCBwbG90LXJlYWR5IEV4Y2VsIHdvcmtib29rLiIiIgogICAgYmFzZSA9IG9zLnBhdGguc3BsaXRleHQob3MucGF0aC5iYXNlbmFtZShjc3ZfcGF0aCkpWzBdCiAgICBzZXRfbmFtZSA9IGJhc2UucmVwbGFjZSgiX2NvbnNvbGlkYXRlZCIsICIiKQoKICAgIGlmIHhsc3hfcGF0aCBpcyBOb25lOgogICAgICAgIHhsc3hfcGF0aCA9IG9zLnBhdGguam9pbihvcy5wYXRoLmRpcm5hbWUoY3N2X3BhdGgpLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAiJXNfUmVzdWx0cy54bHN4IiAlIHNldF9uYW1lKQoKICAgIHByaW50KCJSZWFkaW5nOiAlcyIgJSBjc3ZfcGF0aCkKICAgIG1ldGEsIGhlYWRlciwgcm93cyA9IHBhcnNlX2Nzdihjc3ZfcGF0aCkKICAgIHZhcl9sYWJlbHMgPSBkZXRlY3RfdmFyaWFibGVfbGFiZWxzKGhlYWRlcikKICAgIGlzXzJkID0gbWV0YS5nZXQoIjJEIFN3ZWVwIiwgIk5vIikubG93ZXIoKSA9PSAieWVzIgoKICAgIHByaW50KCIgIFZhcmlhYmxlczogJXMiICUgIiwgIi5qb2luKHZhcl9sYWJlbHMpKQogICAgcHJpbnQoIiAgMkQgU3dlZXA6ICAlcyIgJSBpc18yZCkKICAgIHByaW50KCIgIERhdGEgcm93czogJWQiICUgbGVuKHJvd3MpKQoKICAgIHhfbGFiZWxzLCB5X2xhYmVscywgc3RlcF9vcmRlciwgZGF0YSA9IGJ1aWxkX2RhdGEocm93cywgdmFyX2xhYmVscykKICAgIHByaW50KCIgIFggcG9zaXRpb25zOiAlZCIgJSBsZW4oeF9sYWJlbHMpKQogICAgcHJpbnQoIiAgWSBwb3NpdGlvbnM6ICVkIiAlIGxlbih5X2xhYmVscykpCiAgICBwcmludCgiICBTdGVwczogICAgICAgJWQiICUgbGVuKHN0ZXBfb3JkZXIpKQoKICAgICMgQ2xhc3NpZnkgc3RlcHM6IExvYWRpbmcgKG9kZCBpbmRleCksIFJlbGF4YXRpb24gKGV2ZW4gaW5kZXgpCiAgICBsb2FkaW5nX3N0ZXBzID0gW3MgZm9yIGksIHMgaW4gZW51bWVyYXRlKHN0ZXBfb3JkZXIpIGlmIGkgJSAyID09IDBdCiAgICByZWxheGF0aW9uX3N0ZXBzID0gW3MgZm9yIGksIHMgaW4gZW51bWVyYXRlKHN0ZXBfb3JkZXIpIGlmIGkgJSAyID09IDFdCiAgICBpZiBsZW4oc3RlcF9vcmRlcikgPT0gMToKICAgICAgICBsb2FkaW5nX3N0ZXBzID0gc3RlcF9vcmRlcgogICAgICAgIHJlbGF4YXRpb25fc3RlcHMgPSBbXQoKICAgIHByaW50KCIgIExvYWRpbmcgc3RlcHM6ICAgICAlcyIgJSAiLCAiLmpvaW4obG9hZGluZ19zdGVwcykpCiAgICBwcmludCgiICBSZWxheGF0aW9uIHN0ZXBzOiAgJXMiICUgIiwgIi5qb2luKHJlbGF4YXRpb25fc3RlcHMpKQoKICAgIHdiID0gV29ya2Jvb2soKQogICAgd2IucmVtb3ZlKHdiLmFjdGl2ZSkgICMgUmVtb3ZlIGRlZmF1bHQgZW1wdHkgc2hlZXQKCiAgICAjIFN1bW1hcnkgc2hlZXQKICAgIHdyaXRlX3N1bW1hcnkod2IsIHNldF9uYW1lLCB2YXJfbGFiZWxzLCB4X2xhYmVscywgeV9sYWJlbHMsIHN0ZXBfb3JkZXIsCiAgICAgICAgICAgICAgICAgIGxvYWRpbmdfc3RlcHMsIHJlbGF4YXRpb25fc3RlcHMsIGRhdGEpCgogICAgIyBQZXItdmFyaWFibGUgc2hlZXRzCiAgICB3cml0ZV92YXJpYWJsZV9zaGVldHMod2IsIHZhcl9sYWJlbHMsIHhfbGFiZWxzLCB5X2xhYmVscywgc3RlcF9vcmRlciwKICAgICAgICAgICAgICAgICAgICAgICAgICBsb2FkaW5nX3N0ZXBzLCByZWxheGF0aW9uX3N0ZXBzLCBkYXRhLCBpc18yZCwgc2V0X25hbWUpCgogICAgd2Iuc2F2ZSh4bHN4X3BhdGgpCgogICAgIyBDb3VudCBzaGVldHMKICAgIHNoZWV0X25hbWVzID0gd2Iuc2hlZXRuYW1lcwogICAgcHJpbnQoIlxuQ3JlYXRlZCAlZCBzaGVldHM6IiAlIGxlbihzaGVldF9uYW1lcykpCiAgICBmb3IgbmFtZSBpbiBzaGVldF9uYW1lczoKICAgICAgICBwcmludCgiICAtICVzIiAlIG5hbWUpCiAgICBwcmludCgiXG5TYXZlZDogJXMiICUgeGxzeF9wYXRoKQogICAgcmV0dXJuIHhsc3hfcGF0aAoKCmlmIF9fbmFtZV9fID09ICJfX21haW5fXyI6CiAgICBpZiBsZW4oc3lzLmFyZ3YpID4gMToKICAgICAgICBjc3ZfZmlsZSA9IHN5cy5hcmd2WzFdCiAgICBlbHNlOgogICAgICAgIGNhbmRpZGF0ZXMgPSBbZiBmb3IgZiBpbiBvcy5saXN0ZGlyKCIuIikKICAgICAgICAgICAgICAgICAgICAgIGlmIGYuZW5kc3dpdGgoIl9jb25zb2xpZGF0ZWQuY3N2IildCiAgICAgICAgaWYgbm90IGNhbmRpZGF0ZXM6CiAgICAgICAgICAgIHByaW50KCJObyAqX2NvbnNvbGlkYXRlZC5jc3YgZmlsZXMgZm91bmQgaW4gY3VycmVudCBkaXJlY3RvcnkuIikKICAgICAgICAgICAgc3lzLmV4aXQoMSkKICAgICAgICBpZiBsZW4oY2FuZGlkYXRlcykgPiAxOgogICAgICAgICAgICBwcmludCgiTXVsdGlwbGUgQ1NWIGZpbGVzIGZvdW5kLCBwcm9jZXNzaW5nIGFsbDoiKQogICAgICAgICAgICBmb3IgYyBpbiBjYW5kaWRhdGVzOgogICAgICAgICAgICAgICAgY29udmVydF9jc3ZfdG9fZXhjZWwoYykKICAgICAgICAgICAgc3lzLmV4aXQoMCkKICAgICAgICBjc3ZfZmlsZSA9IGNhbmRpZGF0ZXNbMF0KCiAgICBjb252ZXJ0X2Nzdl90b19leGNlbChjc3ZfZmlsZSkK"
)

# Write converter to output directory
_converter = os.path.join(OUTPUT_DIR, 'format_to_excel.py')
try:
    with open(_converter, 'wb') as _f:
        _f.write(_b64.b64decode(_CONVERTER_B64))
    print('\n  Wrote converter: %s' % _converter)
except Exception as _e:
    print('\n  [WARNING] Could not write converter: %s' % _e)

# Find a working Python 3 and ensure openpyxl is installed
print('\n' + '=' * 70)
print('  GENERATING EXCEL WORKBOOKS ...')
print('=' * 70)

_excel_ok = False
_py_candidates = ['python', 'python3', 'py']

# Also try common Windows Python locations
for _ver in ['313', '312', '311', '310', '39']:
    _py_candidates.append(r'C:\Python%s\python.exe' % _ver)
    _py_candidates.append(os.path.expanduser(r'~\AppData\Local\Programs\Python\Python%s\python.exe' % _ver))

for _py in _py_candidates:
    try:
        # First try to install openpyxl (silently, in case it's missing)
        _sp.call([_py, '-m', 'pip', 'install', '-q', 'openpyxl'],
                 cwd=OUTPUT_DIR, stdout=open(os.devnull, 'w'), stderr=open(os.devnull, 'w'))
        # Then run the converter
        _rc = _sp.call([_py, _converter], cwd=OUTPUT_DIR)
        if _rc == 0:
            _excel_ok = True
            print('  Excel generation successful (using: %s)' % _py)
            break
    except Exception:
        continue

if not _excel_ok:
    print('  [WARNING] Could not auto-generate Excel.')
    print('  To generate manually, install openpyxl and run:')
    print('    pip install openpyxl')
    print('    python format_to_excel.py')
    print('  from: %s' % OUTPUT_DIR)

'''



def _build_extraction_script(data):
    """Build the Abaqus extraction Python script from configuration data."""
    odb_path = data.get("odbPath", "")
    output_dir = data.get("outputDir", "")
    instance_name = data.get("instanceName", "CARTILAGESAMPLE-1")
    model_width = data.get("modelWidth", 1.44)
    model_height = data.get("modelHeight", 1.22)
    element_sets = data.get("elementSets", [])
    regions = data.get("regions", [])
    x_labels = data.get("xLabels", [])
    y_labels = data.get("yLabels", [])
    is_2d = data.get("is2D", False)
    outputs = data.get("outputs", [])
    steps_mode = data.get("stepsMode", "ALL")
    custom_steps = data.get("customSteps", [])
    frame_selection = data.get("frameSelection", "LAST")

    # Format element sets
    elsets_str = "\n".join(
        "    '%s.%s'," % (instance_name, s) if "." not in s else "    '%s'," % s
        for s in element_sets
    )

    # Format regions with xi, yi, labels
    region_lines = []
    for i, r in enumerate(regions):
        xi = r.get("xi", 0)
        yi = r.get("yi", 0)
        x_label = r.get("xLabel", x_labels[xi] if xi < len(x_labels) else "")
        y_label = r.get("yLabel", y_labels[yi] if yi < len(y_labels) else "")
        region_lines.append(
            "    {'name': 'R_%02d', 'x_min': %.6f, 'x_max': %.6f, "
            "'y_min': %.6f, 'y_max': %.6f, 'xi': %d, 'yi': %d, "
            "'x_label': '%s', 'y_label': '%s'},"
            % (i + 1, r["xMin"], r["xMax"], r["yMin"], r["yMax"],
               xi, yi, x_label, y_label)
        )
    regions_str = "\n".join(region_lines)

    # Format outputs
    outputs_str = "\n".join(
        "    ('%s', %s, '%s')," % (
            o["fieldKey"],
            "None" if o.get("compOrInv") is None else "'%s'" % o["compOrInv"],
            o["label"],
        )
        for o in outputs
    )

    # Steps config
    if steps_mode == "ALL":
        steps_config = "'ALL'"
    else:
        steps_config = "[%s]" % ", ".join("'%s'" % s for s in custom_steps)

    return EXTRACTION_SCRIPT_TEMPLATE.format(
        odb_path=odb_path.replace("\\", "\\\\"),
        output_dir=output_dir.replace("\\", "\\\\"),
        instance_name=instance_name,
        model_width=model_width,
        model_height=model_height,
        element_sets=elsets_str,
        regions=regions_str,
        x_labels=repr(x_labels),
        y_labels=repr(y_labels),
        is_2d="True" if is_2d else "False",
        outputs=outputs_str,
        steps_config=steps_config,
        frame_selection=frame_selection,
    )


def _deploy_excel_converter(target_dir):
    """Copy the bundled format_to_excel.py to the target directory."""
    if not target_dir or not os.path.isdir(target_dir):
        return
    dest = os.path.join(target_dir, "format_to_excel.py")
    # Look for the converter relative to this app.py
    candidates = [
        os.path.join(os.path.dirname(__file__), "format_to_excel.py"),
        os.path.join(os.path.dirname(__file__), "..", "format_to_excel.py"),
        os.path.join(os.path.dirname(__file__), "..", "..", "Results", "format_to_excel.py"),
    ]
    source = None
    for c in candidates:
        if os.path.isfile(c):
            source = c
            break

    if source is None:
        # No bundled converter found; skip silently
        return

    # Copy (or update) the converter
    import shutil
    try:
        shutil.copy2(source, dest)
    except Exception:
        pass


@app.route("/api/generate-script", methods=["POST"])
def generate_script():
    """Generate the Abaqus extraction script and save it."""
    data = request.json
    output_dir = data.get("outputDir", "")

    script_content = _build_extraction_script(data)

    # Determine save path
    if output_dir and os.path.isdir(output_dir):
        script_path = os.path.join(output_dir, "extract_region_sweep.py")
    else:
        script_path = os.path.join(tempfile.gettempdir(), "extract_region_sweep.py")

    with open(script_path, "w") as f:
        f.write(script_content)

    # Also copy the Excel converter script to the output directory
    _deploy_excel_converter(os.path.dirname(script_path))

    return jsonify({
        "scriptPath": script_path,
        "scriptContent": script_content,
    })


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================

def _stream_output(proc, output_dir=None):
    """Read stdout/stderr from process and put into log queue."""
    global _run_status

    # Abaqus writes license info, Intel MPI messages, etc. to stderr
    # These are not actual errors, so we only tag lines that look like real errors
    ERROR_KEYWORDS = ["error", "traceback", "exception", "syntaxerror", "nameerror",
                      "typeerror", "valueerror", "keyerror", "importerror", "runtimeerror"]

    def read_stream(stream, is_stderr):
        for line in iter(stream.readline, ""):
            if line:
                text = line.rstrip("\n\r")
                if is_stderr:
                    lower = text.lower()
                    is_real_error = any(kw in lower for kw in ERROR_KEYWORDS)
                    if is_real_error:
                        _log_queue.put("[ERROR] " + text)
                    else:
                        _log_queue.put(text)
                else:
                    _log_queue.put(text)
        stream.close()

    t_out = threading.Thread(target=read_stream, args=(proc.stdout, False))
    t_err = threading.Thread(target=read_stream, args=(proc.stderr, True))
    t_out.daemon = True
    t_err.daemon = True
    t_out.start()
    t_err.start()

    proc.wait()
    t_out.join(timeout=5)
    t_err.join(timeout=5)

    if proc.returncode == 0:
        _run_status = "completed"
        _log_queue.put("[SYSTEM] Extraction completed successfully.")
        # Auto-generate formatted Excel workbook
        if output_dir and os.path.isdir(output_dir):
            _log_queue.put("[SYSTEM] Generating formatted Excel workbook...")
            try:
                result = _post_process_to_excel(output_dir)
                if "error" in result:
                    _log_queue.put("[SYSTEM] Excel generation failed: %s" % result["error"])
                else:
                    for ep in result.get("excelFiles", []):
                        _log_queue.put("[SYSTEM] Excel saved: %s" % ep)
            except Exception as e:
                _log_queue.put("[SYSTEM] Excel generation error: %s" % str(e))
    else:
        _run_status = "failed"
        _log_queue.put("[SYSTEM] Process exited with code %d" % proc.returncode)


@app.route("/api/run-script", methods=["POST"])
def run_script():
    """Execute the generated script via Abaqus subprocess."""
    global _process, _run_start_time, _run_status

    with _process_lock:
        if _process is not None and _process.poll() is None:
            return jsonify({"error": "A script is already running."}), 400

    data = request.json
    script_path = data.get("scriptPath", "")
    abaqus_cmd = data.get("abaqusPath", "abaqus")
    exec_mode = data.get("execMode", "noGUI")  # noGUI, python, cae
    working_dir = data.get("workingDir", "")

    if not script_path or not os.path.isfile(script_path):
        return jsonify({"error": "Script file not found: %s" % script_path}), 400

    if not working_dir:
        working_dir = os.path.dirname(script_path)
    # Ensure working directory exists and is valid
    if not working_dir or not os.path.isdir(working_dir):
        working_dir = tempfile.gettempdir()

    # Clear log queue
    while not _log_queue.empty():
        try:
            _log_queue.get_nowait()
        except queue.Empty:
            break

    # Build command
    if exec_mode == "noGUI":
        cmd = [abaqus_cmd, "cae", "noGUI=%s" % script_path]
    elif exec_mode == "python":
        cmd = [abaqus_cmd, "python", script_path]
    elif exec_mode == "cae":
        # Just launch CAE
        cmd = [abaqus_cmd, "cae"]
    else:
        return jsonify({"error": "Unknown execution mode: %s" % exec_mode}), 400

    try:
        flags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
        # On Windows, .bat files must be run via shell with a string command
        use_shell = sys.platform == "win32" and abaqus_cmd.lower().endswith(".bat")
        if use_shell:
            cmd = subprocess.list2cmdline(cmd)
        _process = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            text=True, cwd=working_dir, creationflags=flags,
            bufsize=1, shell=use_shell,
        )
        _run_start_time = time.time()
        _run_status = "running"

        _log_queue.put("[SYSTEM] Started: %s" % (cmd if isinstance(cmd, str) else " ".join(cmd)))
        _log_queue.put("[SYSTEM] Working directory: %s" % working_dir)

        # Start output streaming thread (pass output_dir for auto Excel generation)
        t = threading.Thread(target=_stream_output, args=(_process, working_dir))
        t.daemon = True
        t.start()

        return jsonify({
            "status": "running",
            "pid": _process.pid,
            "command": " ".join(cmd),
        })

    except FileNotFoundError:
        return jsonify({"error": "Abaqus command not found: %s" % abaqus_cmd}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/run-status")
def run_status():
    """SSE endpoint for real-time log streaming."""
    def generate():
        while True:
            try:
                line = _log_queue.get(timeout=1)
                yield "data: %s\n\n" % json.dumps({"type": "log", "message": line})
            except queue.Empty:
                # Send heartbeat
                elapsed = time.time() - _run_start_time if _run_start_time else 0
                yield "data: %s\n\n" % json.dumps({
                    "type": "heartbeat",
                    "status": _run_status,
                    "elapsed": round(elapsed, 1),
                })
                if _run_status in ("completed", "failed", "cancelled"):
                    yield "data: %s\n\n" % json.dumps({
                        "type": "done",
                        "status": _run_status,
                        "elapsed": round(elapsed, 1),
                    })
                    break

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/stop-run", methods=["POST"])
def stop_run():
    """Kill the running subprocess."""
    global _run_status

    with _process_lock:
        if _process is not None and _process.poll() is None:
            _process.kill()
            _run_status = "cancelled"
            _log_queue.put("[SYSTEM] Process killed by user.")
            return jsonify({"status": "cancelled"})
        else:
            return jsonify({"status": "not_running"})


# ============================================================================
# POST-PROCESSING: Consolidated CSV -> Excel
# ============================================================================

def _post_process_to_excel(output_dir):
    """
    Read consolidated CSVs and metadata from output_dir.
    Creates plot-ready Excel files organized by Loading vs Relaxation phases.

    Steps are classified as:
      - Loading:     1st, 3rd, 5th, 7th, 9th, ... (odd-indexed steps)
      - Relaxation:  2nd, 4th, 6th, 8th, 10th, ... (even-indexed steps)

    Sheet structure for 1D sweep (e.g. X sweep):
      "S22_Max_Loading"  => rows=X positions, columns=Loading step names
      "S22_Min_Loading"  => rows=X positions, columns=Loading step names
      "S22_Max_Relax"    => rows=X positions, columns=Relaxation step names
      "S22_Min_Relax"    => rows=X positions, columns=Relaxation step names

    For 2D sweep: one sheet per phase+stat with stacked blocks per step.

    Also creates a "Summary" sheet with peak values per step per variable.
    """
    import csv

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    meta_path = os.path.join(output_dir, "_extraction_metadata.json")
    if not os.path.isfile(meta_path):
        return {"error": "Metadata file not found. Run extraction first."}

    with open(meta_path, "r") as f:
        meta = json.load(f)

    header_labels = meta["header_labels"]
    x_labels = meta["x_labels"]
    y_labels = meta["y_labels"]
    is_2d = meta["is_2d"]
    step_names = meta["step_names"]
    element_sets = meta["element_sets"]

    # Classify steps into Loading (odd: 1st, 3rd, ...) and Relaxation (even: 2nd, 4th, ...)
    loading_steps = [s for i, s in enumerate(step_names) if i % 2 == 0]
    relaxation_steps = [s for i, s in enumerate(step_names) if i % 2 == 1]

    # If only 1 step, treat it as loading only
    if len(step_names) == 1:
        loading_steps = step_names
        relaxation_steps = []

    # Style definitions
    header_font = Font(bold=True, size=10)
    header_fill_dark = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    loading_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    loading_font = Font(bold=True, size=10, color="90CDF4")
    relax_fill = PatternFill(start_color="322659", end_color="322659", fill_type="solid")
    relax_font = Font(bold=True, size=10, color="D6BCFA")
    summary_header_fill = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")
    summary_header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    pos_label_font = Font(bold=False, size=10, color="A0AEC0")
    pos_label_fill = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")

    def _auto_fit(ws):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 22)

    def _style_header_row(ws, row_num, num_cols, fill=None, font=None):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = font or header_font_white
            cell.fill = fill or header_fill_dark
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    def _write_value(ws, row_num, col_num, val_str):
        try:
            val = float(val_str)
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.number_format = '0.0000E+00'
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        except (ValueError, TypeError):
            cell = ws.cell(row=row_num, column=col_num, value=val_str)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    def _write_pos_label(ws, row_num, col_num, value):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.font = pos_label_font
        cell.fill = pos_label_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    def _safe_sheet_name(name, max_len=31):
        """Truncate and sanitize sheet name to Excel's 31-char limit."""
        # Remove invalid chars
        for ch in ['\\', '/', '*', '?', ':', '[', ']']:
            name = name.replace(ch, '')
        return name[:max_len]

    excel_files = []

    for elset_name in element_sets:
        set_short = elset_name.split(".")[-1] if "." in elset_name else elset_name
        csv_path = os.path.join(output_dir, "%s_consolidated.csv" % set_short)

        if not os.path.isfile(csv_path):
            continue

        # Read CSV data
        rows = []
        with open(csv_path, "r") as f:
            reader = csv.reader(f)
            header_row = None
            for row in reader:
                if not row or row[0].startswith("#"):
                    continue
                if header_row is None:
                    header_row = row
                    continue
                rows.append(row)

        if not header_row or not rows:
            continue

        col_map = {name: idx for idx, name in enumerate(header_row)}

        # Build data: data[(step, xi, yi)][var_label] = (max_val, min_val)
        data = {}
        for row in rows:
            step = row[col_map["Step"]]
            xi = int(row[col_map["Xi"]])
            yi = int(row[col_map["Yi"]])
            key = (step, xi, yi)
            if key not in data:
                data[key] = {}
            for label in header_labels:
                max_col = "%s_Max" % label
                min_col = "%s_Min" % label
                max_val = row[col_map[max_col]] if max_col in col_map else "N/A"
                min_val = row[col_map[min_col]] if min_col in col_map else "N/A"
                data[key][label] = (max_val, min_val)

        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        # =============================================
        # SUMMARY SHEET
        # =============================================
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.cell(row=1, column=1, value="Element Set: %s" % set_short)
        ws_sum.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
        ws_sum.cell(row=2, column=1, value="Steps: %d (%d Loading + %d Relaxation)" % (
            len(step_names), len(loading_steps), len(relaxation_steps)))
        ws_sum.cell(row=2, column=1).font = Font(size=10, color="A0AEC0")
        ws_sum.cell(row=3, column=1, value="Regions: %d" % (len(x_labels) * max(len(y_labels), 1)))
        ws_sum.cell(row=3, column=1).font = Font(size=10, color="A0AEC0")

        # Summary table: peak values across all regions per step per variable
        sum_row = 5
        ws_sum.cell(row=sum_row, column=1, value="Step")
        ws_sum.cell(row=sum_row, column=2, value="Phase")
        col_idx = 3
        for label in header_labels:
            ws_sum.cell(row=sum_row, column=col_idx, value="%s Max" % label)
            ws_sum.cell(row=sum_row, column=col_idx + 1, value="%s Min" % label)
            col_idx += 2
        _style_header_row(ws_sum, sum_row, col_idx - 1, summary_header_fill, summary_header_font)

        for step_idx, step in enumerate(step_names):
            r = sum_row + step_idx + 1
            phase = "Loading" if step in loading_steps else "Relaxation"
            ws_sum.cell(row=r, column=1, value=step).border = thin_border
            phase_cell = ws_sum.cell(row=r, column=2, value=phase)
            phase_cell.border = thin_border
            if phase == "Loading":
                phase_cell.font = loading_font
                phase_cell.fill = loading_fill
            else:
                phase_cell.font = relax_font
                phase_cell.fill = relax_fill

            col_idx = 3
            for label in header_labels:
                # Find peak across all regions for this step
                peak_max = None
                peak_min = None
                all_xi = range(len(x_labels)) if x_labels else [0]
                all_yi = range(len(y_labels)) if y_labels else [0]
                for xi in all_xi:
                    for yi in all_yi:
                        key = (step, xi, yi)
                        if key in data and label in data[key]:
                            try:
                                mv = float(data[key][label][0])
                                if peak_max is None or mv > peak_max:
                                    peak_max = mv
                            except (ValueError, TypeError):
                                pass
                            try:
                                nv = float(data[key][label][1])
                                if peak_min is None or nv < peak_min:
                                    peak_min = nv
                            except (ValueError, TypeError):
                                pass

                if peak_max is not None:
                    _write_value(ws_sum, r, col_idx, "%.8e" % peak_max)
                else:
                    ws_sum.cell(row=r, column=col_idx, value="N/A").border = thin_border
                if peak_min is not None:
                    _write_value(ws_sum, r, col_idx + 1, "%.8e" % peak_min)
                else:
                    ws_sum.cell(row=r, column=col_idx + 1, value="N/A").border = thin_border
                col_idx += 2

        _auto_fit(ws_sum)

        # =============================================
        # DATA SHEETS: one per variable x stat x phase
        # =============================================
        phases = [("Loading", loading_steps, loading_fill, loading_font),
                  ("Relax", relaxation_steps, relax_fill, relax_font)]

        for label in header_labels:
            for stat_type, stat_idx in [("Max", 0), ("Min", 1)]:
                for phase_name, phase_steps, phase_fill, phase_font in phases:
                    if not phase_steps:
                        continue

                    sheet_name = _safe_sheet_name("%s_%s_%s" % (label, stat_type, phase_name))
                    ws = wb.create_sheet(title=sheet_name)

                    if is_2d:
                        # 2D sweep: stacked blocks, one per step
                        # Each block: rows=Y positions, columns=X positions
                        current_row = 1
                        for step in phase_steps:
                            # Step header
                            ws.cell(row=current_row, column=1,
                                    value="%s  |  %s %s  |  %s" % (step, label, stat_type, phase_name))
                            ws.cell(row=current_row, column=1).font = phase_font
                            ws.cell(row=current_row, column=1).fill = phase_fill
                            ws.merge_cells(start_row=current_row, start_column=1,
                                           end_row=current_row, end_column=len(x_labels) + 1)
                            current_row += 1

                            # Column headers (X labels)
                            ws.cell(row=current_row, column=1, value="Y \\ X")
                            for ci, xl in enumerate(x_labels):
                                ws.cell(row=current_row, column=ci + 2, value=xl)
                            _style_header_row(ws, current_row, len(x_labels) + 1)
                            current_row += 1

                            # Data rows (Y labels)
                            for yi_idx, yl in enumerate(y_labels):
                                _write_pos_label(ws, current_row, 1, yl)
                                for xi_idx in range(len(x_labels)):
                                    key = (step, xi_idx, yi_idx)
                                    if key in data and label in data[key]:
                                        _write_value(ws, current_row, xi_idx + 2,
                                                     data[key][label][stat_idx])
                                current_row += 1

                            current_row += 1  # Blank row between blocks

                    else:
                        # 1D sweep: rows=positions, columns=step names
                        # This is the plot-friendly format!
                        pos_labels = x_labels if len(x_labels) > 1 else y_labels
                        pos_axis = "X" if len(x_labels) > 1 else "Y"

                        # Title row
                        ws.cell(row=1, column=1,
                                value="%s  |  %s %s  |  %s" % (set_short, label, stat_type, phase_name))
                        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
                        ws.merge_cells(start_row=1, start_column=1,
                                       end_row=1, end_column=len(phase_steps) + 1)
                        ws.cell(row=1, column=1).fill = phase_fill

                        # Header row: Position | Step1 | Step2 | ...
                        ws.cell(row=2, column=1, value="Position_%s (mm)" % pos_axis)
                        for si, step in enumerate(phase_steps):
                            ws.cell(row=2, column=si + 2, value=step)
                        _style_header_row(ws, 2, len(phase_steps) + 1, phase_fill, phase_font)

                        # Data rows
                        for ci, pl in enumerate(pos_labels):
                            row_num = ci + 3
                            _write_pos_label(ws, row_num, 1, pl)

                            for si, step in enumerate(phase_steps):
                                if pos_axis == "X":
                                    key = (step, ci, 0)
                                else:
                                    key = (step, 0, ci)

                                if key in data and label in data[key]:
                                    _write_value(ws, row_num, si + 2, data[key][label][stat_idx])

                    _auto_fit(ws)

        # =============================================
        # ALL-STEPS SHEETS (for single-step or custom use)
        # One sheet per variable+stat with ALL steps as columns
        # =============================================
        if len(step_names) > 1:
            for label in header_labels:
                for stat_type, stat_idx in [("Max", 0), ("Min", 1)]:
                    sheet_name = _safe_sheet_name("%s_%s_AllSteps" % (label, stat_type))
                    ws = wb.create_sheet(title=sheet_name)

                    pos_labels = x_labels if len(x_labels) > 1 else y_labels
                    pos_axis = "X" if len(x_labels) > 1 else "Y"

                    if is_2d:
                        # For 2D, create stacked blocks for all steps
                        current_row = 1
                        for step in step_names:
                            step_idx = step_names.index(step)
                            phase = "Loading" if step in loading_steps else "Relaxation"
                            pf = loading_fill if phase == "Loading" else relax_fill
                            pfont = loading_font if phase == "Loading" else relax_font

                            ws.cell(row=current_row, column=1,
                                    value="%s  |  %s %s  |  %s" % (step, label, stat_type, phase))
                            ws.cell(row=current_row, column=1).font = pfont
                            ws.cell(row=current_row, column=1).fill = pf
                            ws.merge_cells(start_row=current_row, start_column=1,
                                           end_row=current_row, end_column=len(x_labels) + 1)
                            current_row += 1

                            ws.cell(row=current_row, column=1, value="Y \\ X")
                            for ci, xl in enumerate(x_labels):
                                ws.cell(row=current_row, column=ci + 2, value=xl)
                            _style_header_row(ws, current_row, len(x_labels) + 1)
                            current_row += 1

                            for yi_idx, yl in enumerate(y_labels):
                                _write_pos_label(ws, current_row, 1, yl)
                                for xi_idx in range(len(x_labels)):
                                    key = (step, xi_idx, yi_idx)
                                    if key in data and label in data[key]:
                                        _write_value(ws, current_row, xi_idx + 2,
                                                     data[key][label][stat_idx])
                                current_row += 1
                            current_row += 1
                    else:
                        # 1D: rows=positions, columns=ALL step names
                        ws.cell(row=1, column=1,
                                value="%s  |  %s %s  |  All Steps" % (set_short, label, stat_type))
                        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
                        ws.merge_cells(start_row=1, start_column=1,
                                       end_row=1, end_column=len(step_names) + 1)
                        ws.cell(row=1, column=1).fill = header_fill_dark

                        # Header with color-coded step names
                        ws.cell(row=2, column=1, value="Position_%s (mm)" % pos_axis)
                        for si, step in enumerate(step_names):
                            cell = ws.cell(row=2, column=si + 2, value=step)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center")
                            if step in loading_steps:
                                cell.font = loading_font
                                cell.fill = loading_fill
                            else:
                                cell.font = relax_font
                                cell.fill = relax_fill

                        # Style position header
                        pos_cell = ws.cell(row=2, column=1)
                        pos_cell.font = header_font_white
                        pos_cell.fill = header_fill_dark
                        pos_cell.border = thin_border

                        # Data rows
                        for ci, pl in enumerate(pos_labels):
                            row_num = ci + 3
                            _write_pos_label(ws, row_num, 1, pl)

                            for si, step in enumerate(step_names):
                                if pos_axis == "X":
                                    key = (step, ci, 0)
                                else:
                                    key = (step, 0, ci)
                                if key in data and label in data[key]:
                                    _write_value(ws, row_num, si + 2, data[key][label][stat_idx])

                    _auto_fit(ws)

        # Save Excel
        excel_filename = "%s_Results.xlsx" % set_short
        excel_path = os.path.join(output_dir, excel_filename)
        wb.save(excel_path)
        excel_files.append(excel_path)

    return {"excelFiles": excel_files, "count": len(excel_files)}


@app.route("/api/post-process-excel", methods=["POST"])
def post_process_excel():
    """Create consolidated Excel files from extraction CSVs."""
    data = request.json
    output_dir = data.get("outputDir", "")

    if not output_dir or not os.path.isdir(output_dir):
        return jsonify({"error": "Output directory not found: %s" % output_dir}), 400

    result = _post_process_to_excel(output_dir)
    if "error" in result:
        return jsonify(result), 500

    return jsonify(result)


@app.route("/api/open-folder", methods=["POST"])
def open_folder():
    """Open a folder in the system file explorer."""
    folder = request.json.get("path", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Directory not found"}), 400

    if sys.platform == "win32":
        os.startfile(folder)
    else:
        subprocess.Popen(["xdg-open", folder])
    return jsonify({"status": "opened"})


@app.route("/assets/<path:filename>")
def serve_assets(filename):
    """Serve static assets with correct MIME types (fixes Windows issue)."""
    mimemap = {".js": "application/javascript", ".css": "text/css",
               ".map": "application/json", ".woff2": "font/woff2"}
    ext = os.path.splitext(filename)[1].lower()
    resp = send_from_directory(os.path.join(app.static_folder, "assets"), filename)
    if ext in mimemap:
        resp.headers["Content-Type"] = mimemap[ext]
    return resp


@app.route("/")
def serve_index():
    """Serve the built frontend."""
    return send_from_directory(app.static_folder, "index.html")


@app.errorhandler(404)
def fallback(e):
    """SPA fallback - serve index.html for non-API routes."""
    return send_from_directory(app.static_folder, "index.html")


if __name__ == "__main__":
    print("Starting Abaqus ODB Extractor on http://localhost:5000")
    app.run(host="127.0.0.1", port=5000, debug=True)
